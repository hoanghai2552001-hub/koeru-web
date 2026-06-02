#!/usr/bin/env python3
"""
KOERU — Import from Excel
==========================
Dùng: python tools/import_excel.py [file.xlsx]
      python tools/import_excel.py            ← tự tìm file mới nhất trong tools/

Đọc file Excel đã sửa và ghi lại vào JS files.
Sau đó tự chạy sync để cập nhật kanji-data.js.

Chỉ cập nhật:
  • Cột "meaning" trong sheet N5/N4/N3/N2/N1
  • Cột "meaning" trong sheet Words (words[].m)
  • Cột "meaning_vi" / "meaning_en" trong sheet Radicals
"""

import re, sys, json, shutil
from pathlib import Path
from datetime import datetime
import openpyxl

PROJECT = Path(__file__).parent.parent
JS      = PROJECT / "js"
LEVELS  = ["n5", "n4", "n3", "n2", "n1"]

# ── Helpers ───────────────────────────────────────────────────────────────


def detect_format(wb) -> str:
    """Nhận diện format: 'multi' (nhiều sheet N5/N4/...) hoặc 'single' (1 sheet tổng)."""
    sheets = wb.sheetnames
    if any(s in sheets for s in ["N5","N4","N3","N2","N1"]):
        return "multi"
    return "single"


def import_single_sheet(wb) -> int:
    """Import format 1 sheet tổng (export từ game panel)."""
    import re, shutil
    ws = wb.active
    headers = [c.value for c in ws[1]]
    try:
        ki = headers.index("kanji")
        mi = headers.index("meaning")
    except ValueError:
        print("    ⚠  Không tìm thấy cột kanji/meaning")
        return 0

    updates = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[ki]:
            continue
        k = str(row[ki]).strip()
        m = str(row[mi]).strip() if row[mi] else ""
        if k and m:
            updates[k] = m

    total = 0
    for level in LEVELS:
        p = JS / f"kanji-data-{level}.js"
        if not p.exists():
            continue
        text = p.read_text(encoding="utf-8")
        changed = 0

        def rep(match, _upd=updates):
            nonlocal changed
            k = match.group(1)
            if k not in _upd:
                return match.group(0)
            new_m = _upd[k].replace('"', '\\"')
            old = match.group(0)
            new = re.sub(r'meaning:".*?"', f'meaning:"{new_m}"', old, count=1)
            if new != old:
                changed += 1
            return new

        new_text = re.sub(r'\{kanji:"(.)".+', rep, text)
        if changed:
            shutil.copy2(p, str(p) + ".xlsbak")
            p.write_text(new_text, encoding="utf-8")
            print(f"  ✅  kanji-data-{level}.js: {changed} meanings cập nhật")
            total += changed
        else:
            print(f"  ○   kanji-data-{level}.js: không có thay đổi")
    return total


def find_latest_excel() -> Path | None:
    files = sorted((PROJECT / "tools").glob("kanji_data_*.xlsx"),
                   key=lambda f: f.stat().st_mtime, reverse=True)
    return files[0] if files else None


def load_js(level: str) -> str:
    path = JS / f"kanji-data-{level}.js"
    return path.read_text(encoding="utf-8") if path.exists() else ""


def save_js(level: str, content: str):
    path = JS / f"kanji-data-{level}.js"
    shutil.copy2(path, str(path) + ".xlsbak")
    path.write_text(content, encoding="utf-8")


# ── Import Level sheets (sửa meaning của kanji) ───────────────────────────

def import_level_sheet(ws) -> dict[str, str]:
    """Đọc sheet N5/N4/... → {kanji_char: new_meaning}"""
    headers = [c.value for c in ws[1]]
    try:
        ki = headers.index("kanji")
        mi = headers.index("meaning")
    except ValueError:
        return {}

    changes = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[ki]:
            continue
        kanji = str(row[ki]).strip()
        meaning = str(row[mi]).strip() if row[mi] is not None else ""
        if kanji and meaning:
            changes[kanji] = meaning
    return changes


def apply_kanji_meanings(level: str, changes: dict[str, str]) -> int:
    """Áp changes vào kanji-data-{level}.js, cập nhật field meaning."""
    content = load_js(level)
    if not content:
        return 0
    updated = 0

    def replacer(m):
        nonlocal updated
        kanji = m.group(1)
        if kanji in changes:
            new_m = changes[kanji].replace('"', '\\"')
            old_line = m.group(0)
            new_line = re.sub(r'meaning:".*?"', f'meaning:"{new_m}"', old_line, count=1)
            if new_line != old_line:
                updated += 1
            return new_line
        return m.group(0)

    # Match từng entry (1 dòng = 1 kanji)
    new_content = re.sub(r'\{kanji:"(.)".+', replacer, content)
    if updated:
        save_js(level, new_content)
    return updated


# ── Import Words sheet (sửa words[].m) ───────────────────────────────────

def import_words_sheet(ws) -> dict[tuple, str]:
    """Đọc sheet Words → {(kanji_char, word, reading): new_meaning}"""
    headers = [c.value for c in ws[1]]
    try:
        ki = headers.index("kanji")
        wi = headers.index("word")
        ri = headers.index("reading")
        mi = headers.index("meaning")
    except ValueError:
        return {}

    changes = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[ki]:
            continue
        kanji   = str(row[ki]).strip()
        word    = str(row[wi]).strip() if row[wi] else ""
        reading = str(row[ri]).strip() if row[ri] else ""
        meaning = str(row[mi]).strip() if row[mi] is not None else ""
        if kanji and word and meaning:
            changes[(kanji, word, reading)] = meaning
    return changes


def apply_word_meanings(level: str, changes: dict[tuple, str]) -> int:
    content = load_js(level)
    if not content:
        return 0

    updated = 0
    # Group changes by kanji
    by_kanji: dict[str, list] = {}
    for (kanji, word, reading), meaning in changes.items():
        by_kanji.setdefault(kanji, []).append((word, reading, meaning))

    def entry_replacer(m):
        nonlocal updated
        kanji = m.group(1)
        if kanji not in by_kanji:
            return m.group(0)
        entry = m.group(0)
        for word, reading, new_meaning in by_kanji[kanji]:
            # Match word entry trong words[]
            word_pat = re.compile(
                r'(\{"w":"' + re.escape(word) + r'","r":"' + re.escape(reading) + r'","m":")(.*?)("\})'
            )
            def word_replacer(wm, nm=new_meaning):
                nonlocal updated
                if wm.group(2) != nm:
                    updated += 1
                return wm.group(1) + nm.replace('"', '\\"') + wm.group(3)
            entry = word_pat.sub(word_replacer, entry)
        return entry

    new_content = re.sub(r'\{kanji:"(.)".+', entry_replacer, content)
    if updated:
        save_js(level, new_content)
    return updated


# ── Import Radicals sheet ─────────────────────────────────────────────────

def import_radicals_sheet(ws) -> int:
    map_path = PROJECT / "kanji-map-data.js"
    if not map_path.exists():
        return 0

    headers = [c.value for c in ws[1]]
    # Header labels: "Ký tự", "Tên JP", "Nghĩa VI", "Nghĩa EN"
    try:
        ki  = next(i for i, h in enumerate(headers) if h in ("key", "Ký tự"))
        mvi = next(i for i, h in enumerate(headers) if h in ("meaning_vi", "Nghĩa VI"))
        men = next(i for i, h in enumerate(headers) if h in ("meaning_en", "Nghĩa EN"))
    except StopIteration:
        print("    ⚠  Không tìm thấy cột key/meaning_vi/meaning_en")
        return 0

    changes_vi: dict[str, str] = {}
    changes_en: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[ki]:
            continue
        key = str(row[ki]).strip()
        vi  = str(row[mvi]).strip() if row[mvi] else ""
        en  = str(row[men]).strip() if row[men] else ""
        if key:
            changes_vi[key] = vi
            changes_en[key] = en

    content = map_path.read_text(encoding="utf-8")
    updated = 0

    def rad_replacer(m):
        nonlocal updated
        key = m.group(1)
        if key not in changes_vi:
            return m.group(0)
        line = m.group(0)
        new_vi = changes_vi[key].replace('"', '\\"')
        new_en = changes_en.get(key, "").replace('"', '\\"')
        new_line = re.sub(r'meaning:".*?"', f'meaning:"{new_vi}"', line, count=1)
        new_line = re.sub(r'meaning_en:".*?"', f'meaning_en:"{new_en}"', new_line, count=1)
        if new_line != line:
            updated += 1
        return new_line

    new_content = re.sub(r'\s+"(.+?)":\s*\{[^\n]+\}', rad_replacer, content)
    if updated:
        shutil.copy2(map_path, str(map_path) + ".xlsbak")
        map_path.write_text(new_content, encoding="utf-8")
    return updated


# ── Main ──────────────────────────────────────────────────────────────────

def main():
    # Tìm file Excel
    if len(sys.argv) > 1:
        xlsx_path = Path(sys.argv[1])
        if not xlsx_path.exists():
            xlsx_path = PROJECT / "tools" / sys.argv[1]
    else:
        xlsx_path = find_latest_excel()

    if not xlsx_path or not xlsx_path.exists():
        print("❌ Không tìm thấy file Excel. Dùng: python tools/import_excel.py <file.xlsx>")
        return

    print(f"\n📥 KOERU Import Excel")
    print(f"   File: {xlsx_path.name}\n")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    fmt = detect_format(wb)
    total_updated = 0

    if fmt == "single":
        print("  Format: 1 sheet tổng (game panel export)\n")
        total_updated = import_single_sheet(wb)
    else:
        print("  Format: nhiều sheet (tools/export_excel.py)\n")

        # ── Level sheets → kanji meaning ──
        for level in LEVELS:
            sheet_name = level.upper()
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            changes = import_level_sheet(ws)

            n_kanji = apply_kanji_meanings(level, changes)

            if n_kanji:
                print(f"  ✅  {sheet_name}: {n_kanji} kanji meaning đã cập nhật")
            else:
                print(f"  ○   {sheet_name}: không có thay đổi")
            total_updated += n_kanji

        # ── Words sheet → words[].m ──
        if "Words" in wb.sheetnames:
            word_changes = import_words_sheet(wb["Words"])
            w_total = 0
            for level in LEVELS:
                n = apply_word_meanings(level, word_changes)
                w_total += n
            if w_total:
                print(f"  ✅  Words: {w_total} word meanings đã cập nhật")
            else:
                print(f"  ○   Words: không có thay đổi")
            total_updated += w_total

        # ── Radicals sheet ──
        if "Radicals" in wb.sheetnames:
            n_rad = import_radicals_sheet(wb["Radicals"])
            if n_rad:
                print(f"  ✅  Radicals: {n_rad} bộ thủ đã cập nhật")
            else:
                print(f"  ○   Radicals: không có thay đổi")
            total_updated += n_rad

    # ── Auto build (sync + version bump + study index + quality check) ──
    print()
    if total_updated:
        print(f"  Tổng: {total_updated} thay đổi — đang build...")
        try:
            import subprocess
            result = subprocess.run(
                [sys.executable, str(PROJECT / "tools" / "build.py"), "--quick"],
                cwd=str(PROJECT), capture_output=False
            )
        except Exception as e:
            print(f"  ⚠  Build thủ công: python tools/build.py ({e})\n")
    else:
        print("  ○  Không có thay đổi nào — file JS giữ nguyên\n")


if __name__ == "__main__":
    main()
