"""
gen_final_excel.py — Xuất Excel cuối cùng với nghĩa tiếng Việt đầy đủ
Tích hợp VI dict lớn + detection chuẩn xác + format đẹp
"""
import re
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent.parent
JS  = ROOT / "js"
XLS = ROOT / "input" / "excel"
OUT = ROOT / "output"
OUT.mkdir(exist_ok=True)

# ── Từ Vietnamese không dấu (whitelist để tránh false positive) ───────────────
VIET_WORDS = {
    'xanh','lam','den','do','vang','trang','tim','cam','hong','nau','xam',
    'mot','hai','ba','bon','nam','sau','bay','tam','chin','muoi','tram','ngan',
    'ngay','tuan','thang','nam','gio','phut','giay','som','muon','ngay',
    'nha','cua','ban','ghe','san','tuong','man','man','cua','tang',
    'com','canh','an','uong','nau','mon','bua','sang','trua','toi','khuya',
    'me','bo','anh','chi','em','ong','ba','chau','con','chong','vo',
    'toi','ban','anh','chi','ho','chung','ta','minh',
    'di','den','ve','ra','vao','len','xuong','qua','sang',
    'doc','viet','nghe','noi','nhin','xem','biet','hieu','nho','quen',
    'an','uong','ngu','day','ngoi','dung','chay','boi','hat','mua','ve',
    'lam','lam','hoi','tra','giai','bat','ket','tiep','dung','mua',
    'tot','xau','dep','moi','cu','nho','lon','ngang','dai','ngan',
    'nong','lanh','am','mat','sang','toi','sach','ban','on','vang',
    'nhanh','cham','som','muon','nhieu','it','du','thieu',
    'vui','buon','gian','so','lo','thich','ghet','yeu',
    'rat','qua','kha','cung','just','van','chua','roi','da',
    'va','hay','hoac','nhung','ma','tuy','neu','vi','de',
    'den','tu','trong','ngoai','tren','duoi','giua','ben','canh',
    'day','do','kia','nay','nao','gi','ai','sao','bao',
    'roi','ngay','dung','chinh','deu','moi','nhau','voi',
    'hay','thuong','luc','khi','truoc','sau','dau','cuoi',
    'chiec','quyen','nguoi','lan','cai','mon','bai','to','chu',
    'biet','hoc','day','lop','sach','bai','gio','tiet',
    'phong','san','nha','vien','cong','truong','cho','ngan',
    'xe','may','tau','bus','oto','gia','dinh','phap',
    'tieng','tu','nghia','doc','viet','cau','bai','van',
    'bao','thu','thi','diem','mon','lop','truong','hoc',
    'viec','lam','phong','ban','hop','lich','ke',
    'em trai','em gai','anh trai','chi gai','bo me','ong ba','con trai','con gai',
    'khi','qua','ma','cach','tu','nhu','trong','theo','bang',
}

# Thêm từ có thể là Vietnamese không dấu
VIET_PREFIXES = {'khong','khoa','nguon','ngoai','quan','truoc','phong','duong','cong','viet','thai','trung','nhat','han'}

def is_still_en(s):
    """Kiểm tra xem chuỗi có CÒN tiếng Anh không — tránh false positive"""
    if not isinstance(s, str) or not s.strip(): return False
    # Có dấu tiếng Việt → đã là VI
    viet_chars = 'àáảãạăắặằẳẵâầấẩẫậèéẻẽẹêềếểễệìíỉĩịòóỏõọôồốổỗộơờớởỡợùúủũụưừứửữựỳýỷỹỵđ'
    if any(c in viet_chars for c in s): return False
    # Lấy tất cả các từ Latin
    words = re.findall(r'[a-zA-Z]{3,}', s.lower())
    if not words: return False
    # Nếu tất cả từ đều nằm trong whitelist → đã là VI
    if all(w in VIET_WORDS or w in VIET_PREFIXES for w in words): return False
    # Có từ tiếng Anh thực sự
    EN_MARKERS = {
        'to','the','is','are','was','were','be','been','have','has','had',
        'do','does','did','will','would','should','could','can','may','might',
        'that','this','for','with','from','into','onto','upon','about','above',
        'after','before','under','over','between','through','during','while',
        'and','but','or','not','very','most','more','some','any','all','each',
        'how','why','when','where','what','which','who','whose','whom',
        'food','good','bad','old','new','big','small','high','low',
        'verb','noun','adj','counter','honorific','suffix','prefix',
        'reference','abbr','lit','polite','formal','informal','humble',
        'object','person','things','items','place','time','direction',
        'number','amount','degree','level','period','section','type',
        'making','being','getting','having','doing','going','coming',
        'usually','always','never','often','sometimes','already','still',
        'various','different','other','same','such','these','those',
        'certain','possible','necessary','important','special','general',
    }
    # Nếu có từ EN marker → còn tiếng Anh
    has_en_marker = any(w in EN_MARKERS for w in words)
    # Nếu có từ dài (>5 ký tự) không phải VI → có khả năng EN
    long_words = [w for w in words if len(w) > 5 and w not in VIET_WORDS and w not in VIET_PREFIXES]
    return has_en_marker or len(long_words) > 0


# ── Bảng dịch tổng hợp ────────────────────────────────────────────────────────
from patch_vi_meanings import VI, translate, load_js_words

# Bổ sung thêm từ còn thiếu
VI.update({
    "fall (season)":"mùa thu","spring":"mùa xuân","summer":"mùa hè","winter":"mùa đông",
    "day after tomorrow":"ngày kia (tương lai)","morning (early)":"buổi sáng sớm",
    "coffee shop":"quán cà phê","medicine, drug":"thuốc","(with te-form verb) please do for me":"xin hãy ~ (kèm thể te)",
    "under, below, beneath":"dưới; phía dưới","quiet, calm":"yên tĩnh; bình tĩnh",
    "work, job, occupation, employment":"công việc; việc làm",
    "there, over there, that place":"ở đó; đằng kia","that over there; like that, that way; um...":"cái kia; như vậy; ừm...",
    "pencil":"bút chì","cloudiness":"trời mây; âm u","sugar":"đường",
    "to copy":"sao chép; copy","ticket":"vé","black tea":"trà đen; hồng trà",
    "month":"tháng","suburb":"ngoại ô","trash, garbage":"rác; rác thải",
    "mark, score, grade; point, dot":"điểm; chấm","very hard (as in 'to work hard'), with utmost effort":"hết sức cố gắng",
    "chance, opportunity":"cơ hội","to build":"xây dựng; xây",
    "-- honorific expression for いく, くる, and いる --":"đi/đến/ở (kính ngữ)",
    "futon":"đệm ngủ; chăn bông (Nhật)","important, valuable, serious matter":"quan trọng; đáng quý",
    "a gift; a present":"quà tặng","thief; burglar":"kẻ trộm; tên trộm",
    "made in ~":"sản xuất tại ~","sun, sunshine, day":"mặt trời; ánh nắng; ngày",
    "counter for houses":"tòa ~ (nhà)",
    "really, (is that) so; yes, right":"thật vậy; ừ; đúng vậy",
    "report":"báo cáo; bản báo cáo","leaf":"chiếc lá",
    "section manager":"trưởng phòng; trưởng bộ phận",
    "Africa":"châu Phi","month (of year)":"tháng (trong năm)",
    "to snap, to break; to bend":"bẻ; gãy; uốn cong",
    "to be (polite), to exist":"có mặt; tồn tại (kính ngữ)",
    "~ meeting":"buổi họp ~; hội ~",
    "to cease, to stop":"ngừng lại; chấm dứt",
    "surface; front; outside":"bề mặt; mặt trước; bên ngoài",
    "exercise":"tập thể dục; luyện tập",
    "barber's (shop)":"tiệm cắt tóc nam",
    "pronunciation":"phát âm",
    "motorcycle (lit: auto-bi(ke))":"xe máy; mô tô",
    "as I thought, absolutely":"đúng như tôi nghĩ; tất nhiên",
    "stereo":"dàn âm thanh stereo",
    "text; text book":"giáo trình; sách giáo khoa",
    "distinction, different":"sự khác biệt; khác nhau",
    "worry, concern":"lo lắng; mối lo ngại",
    "soft (in reference to texture), tender":"mềm; mềm mại",
    "useless, no good, hopeless":"vô ích; không được; vô vọng",
    "yes (informal), all right (ok)":"ừ; ok (thân mật)",
    "discourtesy, impoliteness; Excuse me":"vô lễ; Xin lỗi",
    "New Year, New Year's Day":"Năm Mới; Ngày đầu năm",
    "mostly, almost":"hầu hết; gần như",
    "gasoline, petrol":"xăng",
    "preparation of lessons (for class)":"chuẩn bị bài học trước",
    "college student, university student":"sinh viên đại học",
    "to bake, to grill":"nướng; nướng lửa","rooftop":"sân thượng; mái nhà",
    "intention, plan":"dự định; ý định",
    "spirit, mood":"tinh thần; tâm trạng",
    "intention":"dự định","mood":"tâm trạng",
    "place where things are sold":"nơi bán hàng; gian hàng",
    "gradually":"dần dần; từ từ","by degrees":"từng bước",
    "It's ok (all right); No need to worry; Everything is under control":"ổn; không sao; yên tâm",
    "great number of people":"rất nhiều người; đám đông",
    "counter for occurrences (~ times)":"~ lần (đếm số lần)",
    "counter for stories (floors) of a building":"~ tầng (tòa nhà)",
    "counter for small items (e.g., fruits, cups)":"~ cái (đồ nhỏ)",
    "counter for books":"~ quyển (sách)","counter for vehicles; machines":"~ chiếc (xe, máy)",
    "counter for people":"~ người","counter for cupfuls":"~ ly; ~ chén",
    "(number of) months":"(số) tháng","~ weeks":"~ tuần",
    "~ hours":"~ giờ (thời gian)","~ years old":"~ tuổi",
    "~ o'clock (time)":"~ giờ (đồng hồ)","at the time of ~":"vào lúc ~; khi ~",
    "~ day of the month, for ~ days":"ngày ~ trong tháng; ~ ngày",
    "~ years":"~ năm","year, age":"năm; tuổi",
    "five days; fifth day of the month":"năm ngày; ngày mùng năm",
    "seven days; seventh day (of the month)":"bảy ngày; ngày mùng bảy",
    "nine days; ninth day of the month":"chín ngày; ngày mùng chín",
    "ten days; tenth day of the month":"mười ngày; ngày mùng mười",
    "month of year":"tháng trong năm",
    "this week":"tuần này","tonight, this evening":"tối nay",
    "year after next":"năm kia (tương lai)","yesterday":"hôm qua",
    "20 years old":"20 tuổi; hai mươi tuổi",
    # N4 vocab phổ biến chưa có
    "important":"quan trọng","valuable":"quý giá","serious matter":"vấn đề nghiêm trọng",
    "to introduce":"giới thiệu","to get married":"kết hôn",
    "to move (house)":"dọn nhà; chuyển chỗ ở",
    "to decide":"quyết định","to choose":"chọn","to cancel":"hủy",
    "to reserve":"đặt chỗ; đặt trước","to pay":"thanh toán; trả tiền",
    "to earn":"kiếm tiền","to save (money)":"tiết kiệm",
    "to repair":"sửa chữa","to use, to utilize":"dùng; sử dụng",
    "to increase":"tăng; tăng lên","to decrease":"giảm; giảm xuống",
    "to win":"thắng","to lose (a game)":"thua",
    "to pass (an exam)":"đậu; qua kỳ thi",
    "to fail (an exam)":"trượt; rớt kỳ thi",
    "to apply for":"nộp đơn; đăng ký",
    "to contact":"liên lạc","to present":"trình bày",
    "to compare":"so sánh","to check":"kiểm tra",
    "to count":"đếm","to measure":"đo",
    "to warm up":"làm nóng","to cool down":"làm lạnh; nguội",
    "to breathe":"thở","to hurt, to ache":"đau",
    "to be surprised":"ngạc nhiên","to be worried":"lo lắng",
    "to be glad, to be happy":"vui; hạnh phúc","to be sad":"buồn",
    "to be angry":"tức giận","to be afraid":"sợ hãi",
    "to be relieved":"nhẹ nhõm","to be embarrassed":"ngượng; xấu hổ",
    "possibility":"khả năng; có thể","impossible":"không thể",
    "possible":"có thể","necessary":"cần thiết",
    "safe":"an toàn","dangerous":"nguy hiểm",
    "careful":"cẩn thận","careless":"cẩu thả; bất cẩn",
    "polite":"lịch sự","impolite":"vô lễ","kind":"tốt bụng",
    "strict":"nghiêm khắc","honest":"trung thực",
    "patient":"kiên nhẫn","active":"năng động","lazy":"lười biếng",
    "shy":"nhút nhát","cheerful":"vui vẻ","clever":"thông minh",
    "handsome":"đẹp trai","beautiful":"đẹp","ugly":"xấu",
    "young":"trẻ","old (person)":"già","tall":"cao",
    "short (person)":"lùn; thấp","fat":"béo","thin (person)":"gầy",
    "healthy":"khỏe mạnh","sick":"bệnh; ốm",
    "busy":"bận rộn","free (time)":"rảnh","expensive":"đắt tiền","cheap":"rẻ",
    "large, big":"to; lớn","small, little":"nhỏ","long":"dài",
    "short (length)":"ngắn","wide":"rộng","narrow":"hẹp",
    "high":"cao","low":"thấp","heavy":"nặng","light":"nhẹ",
    "fast, quick":"nhanh","slow":"chậm","early":"sớm","late":"muộn; trễ",
    "new":"mới","old":"cũ","hot (weather)":"nóng","cold (weather)":"lạnh",
    "warm":"ấm","cool":"mát","good":"tốt","bad":"xấu; tồi",
    "right, correct":"đúng","wrong":"sai","interesting":"thú vị",
    "boring":"nhàm chán","difficult":"khó","easy":"dễ",
    "clean":"sạch","dirty":"bẩn","quiet":"yên tĩnh",
    "bright":"sáng","dark":"tối","near":"gần","far":"xa",
    "convenient":"tiện lợi","inconvenient":"bất tiện",
    "important":"quan trọng","special":"đặc biệt","ordinary":"bình thường",
    "enough":"đủ","suitable":"phù hợp","popular":"phổ biến",
    "famous":"nổi tiếng","international":"quốc tế","public":"công cộng",
    "modern":"hiện đại","traditional":"truyền thống",
    "professional":"chuyên nghiệp","original":"gốc; bản gốc",
    "real":"thật; thực","fake":"giả","certain, sure":"chắc chắn",
    "so, therefore":"vì vậy; do đó",
    "also":"cũng","even":"thậm chí","only":"chỉ","about":"về; khoảng",
    "between":"giữa","among":"trong số","except":"ngoại trừ",
    "including":"bao gồm","according to":"theo","against":"chống lại",
    "with":"với","without":"không có","until":"cho đến","since":"kể từ",
    "during":"trong suốt","before":"trước","after":"sau",
    "because":"vì; bởi vì","if":"nếu","although":"mặc dù","while":"trong khi",
})


def write_excel_output(out_path, level, vocab_df, kanji_df, word_col, reading_col, meaning_en_col):
    wb = Workbook()

    HDR_COLOR  = "1F3864"
    HDR2_COLOR = "2E4057"
    ROW_ALT    = "EBF0F7"
    WARN_COLOR = "FFF2CC"
    WARN_FONT  = "CC0000"

    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_style(cell, bg=HDR_COLOR):
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    def data_style(cell, alt=False, warn=False):
        cell.font = Font(name="Arial", size=10,
                         color=WARN_FONT if warn else "000000")
        cell.fill = PatternFill("solid", start_color=WARN_COLOR if warn else (ROW_ALT if alt else "FFFFFF"))
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = border

    # ── Summary ────────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    total   = len(vocab_df)
    vi_done = (~vocab_df['Cần kiểm tra']).sum()

    rows = [
        ["KOERU JLPT Database", f"{level} — Vocab & Kanji (chuẩn hoá tiếng Việt)"],
        ["Level", level],
        ["Số từ vựng", total],
        ["Đã Việt hoá", int(vi_done)],
        ["Tỷ lệ", f"{100*vi_done/total:.1f}%"],
        ["Cần review thêm", int(total - vi_done)],
        ["Số kanji", len(kanji_df)],
        ["Nguồn", "open-anki-jlpt-decks + KANJIDIC + KOERU DB"],
        ["Cập nhật", "2026"],
        ["Lưu ý", "Cột 'Nghĩa VI' màu đỏ = cần kiểm tra lại thủ công"],
    ]
    for i, (k, v) in enumerate(rows, 1):
        c1 = ws.cell(row=i, column=1, value=k)
        c2 = ws.cell(row=i, column=2, value=v)
        if i == 1:
            c1.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
            c1.fill = c2.fill = PatternFill("solid", start_color=HDR_COLOR)
            c2.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        else:
            for c in [c1, c2]:
                c.font = Font(name="Arial", size=10, bold=(i % 2 == 0))
                c.fill = PatternFill("solid", start_color=ROW_ALT if i % 2 == 0 else "FFFFFF")
        c1.alignment = c2.alignment = Alignment(vertical="center", wrap_text=True)
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 55
    for i in range(1, len(rows)+1):
        ws.row_dimensions[i].height = 22

    # ── Vocab sheet ────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet(f"Vocab_{level}")
    hanviet_col = 'Hán Việt (theo từng chữ)' if level == 'N4' else 'Hán Việt'
    out_cols = [c for c in ['STT', word_col, reading_col, 'Nghĩa VI', meaning_en_col,
                             'Hán tự trong từ', hanviet_col, 'Cần kiểm tra', 'Nguồn']
                if c in vocab_df.columns]

    for j, col in enumerate(out_cols, 1):
        hdr_style(ws2.cell(row=1, column=j, value=col))

    for i, (_, row) in enumerate(vocab_df[out_cols].iterrows(), 2):
        warn = bool(row.get('Cần kiểm tra', False))
        alt  = i % 2 == 0
        for j, val in enumerate(row, 1):
            cell = ws2.cell(row=i, column=j, value=val if pd.notna(val) else '')
            data_style(cell, alt=alt, warn=(warn and j == out_cols.index('Nghĩa VI')+1))

    col_w = {word_col: 18, reading_col: 16, 'Nghĩa VI': 38, meaning_en_col: 38,
             'Hán tự trong từ': 14, hanviet_col: 22, 'STT': 6,
             'Cần kiểm tra': 14, 'Nguồn': 20}
    for j, col in enumerate(out_cols, 1):
        ws2.column_dimensions[get_column_letter(j)].width = col_w.get(col, 16)
    ws2.row_dimensions[1].height = 30
    ws2.freeze_panes = 'A2'

    # ── Kanji sheet ────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet(f"Kanji_{level}")
    for j, col in enumerate(kanji_df.columns, 1):
        hdr_style(ws3.cell(row=1, column=j, value=col), bg=HDR2_COLOR)
    for i, (_, row) in enumerate(kanji_df.iterrows(), 2):
        for j, val in enumerate(row, 1):
            cell = ws3.cell(row=i, column=j, value=val if pd.notna(val) else '')
            data_style(cell, alt=(i % 2 == 0))
    for j in range(1, len(kanji_df.columns)+1):
        ws3.column_dimensions[get_column_letter(j)].width = 20
    ws3.row_dimensions[1].height = 30
    ws3.freeze_panes = 'A2'

    wb.save(out_path)
    print(f"✅ Đã xuất: {out_path.name}")
    return vi_done, total


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == '__main__':
    js_map = load_js_words()

    configs = [
        ('N5', XLS/"KOERU_JLPT_N5_Vocab_Kanji_OnKun_HanViet.xlsx",
         'Vocab_N5', 'Kanji_N5', 'Từ vựng / Expression', 'Cách đọc / Reading', 'Nghĩa nguồn EN'),
        ('N4', XLS/"KOERU_JLPT_N4_Vocab_Kanji_OnKun_HanViet.xlsx",
         'Vocab_N4', 'Kanji_N4', 'Từ vựng', 'Cách đọc', 'Nghĩa EN nguồn'),
    ]

    for level, src, vocab_sh, kanji_sh, word_col, read_col, en_col in configs:
        print(f"\n{'='*55}\n{level}\n{'='*55}")
        vocab_df = pd.read_excel(src, sheet_name=vocab_sh)
        kanji_df = pd.read_excel(src, sheet_name=kanji_sh)

        vocab_df['Nghĩa VI'] = vocab_df.apply(
            lambda r: translate(r[en_col], js_map, r[word_col]), axis=1
        )
        vocab_df['Cần kiểm tra'] = vocab_df['Nghĩa VI'].apply(is_still_en)

        vi_done, total = write_excel_output(
            OUT / f"KOERU_JLPT_{level}_Updated.xlsx",
            level, vocab_df, kanji_df, word_col, read_col, en_col
        )
        pct = 100 * vi_done / total
        print(f"   Coverage: {vi_done}/{total} ({pct:.1f}%) | còn {total-vi_done} cần review")

    print("\n" + "="*55)
    print("BƯỚC CUỐI: Rebuild JS schema")
    print("="*55)
    # Rebuild combined (schema đã được clean ở bước trước)
    parts = []
    for lv in ['n5','n4','n3','n2','n1']:
        c = (JS/f"kanji-data-{lv}.js").read_text(encoding='utf-8')
        start = c.index('['); end = c.rindex(']') + 1
        inner = c[start+1:end-1].strip().rstrip(',')
        if inner: parts.append(inner)
    combined = 'window.KANJI_DATA = [\n' + ',\n'.join(parts) + '\n];\n'
    (JS/'kanji-data.js').write_text(combined, encoding='utf-8')
    print("✅ kanji-data.js rebuilt")
    print("\n✅ HOÀN TẤT!")
