"""
patch_n3_excel.py
- Đọc 8 file Excel N3
- Dịch nghĩa tiếng Anh → tiếng Việt
- Cập nhật words[] trong kanji-data-n3.js nếu thiếu
- Xuất file Excel tổng hợp: output/KOERU_JLPT_N3_Updated.xlsx
"""
import pandas as pd
import re
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys
sys.path.insert(0, str(Path(__file__).parent))
from vi_n3_translations import VI_N3

BASE = Path(__file__).parent.parent
INPUT = BASE / "input" / "excel"
JS = BASE / "js"
OUTPUT = BASE / "output"
OUTPUT.mkdir(exist_ok=True)

VIET_CHARS = set('àáảãạăắặằẳẵâấầẩẫậđèéẻẽẹêếềệểễìíỉĩịòóỏõọôốồổỗộơớờởỡợùúủũụưứừửữựỳýỷỹỵÀÁẢÃẠĂẮẶẰẲẴÂẤẦẨẪẬĐÈÉẺẼẸÊẾỀỆỂỄÌÍỈĨỊÒÓỎÕỌÔỐỒỔỖỘƠỚỜỞỠỢÙÚỦŨỤƯỨỪỬỮỰỲÝỶỸỴ')

def is_english(s):
    if not s or str(s).strip() in ('nan', ''):
        return True
    s = str(s)
    if any(c in VIET_CHARS for c in s):
        return False
    # Has Japanese → not English
    if any('　' <= c <= '鿿' or '＀' <= c <= '￯' for c in s):
        return False
    viet_words = {'hay', 'co', 'mot', 'the', 'la', 'va', 'khong', 'cung', 'dang',
                  'tren', 'duoi', 'trong', 'ngoai', 'tu', 'so', 'den', 'vi', 'cua',
                  'nhu', 'khi', 'neu', 'sau', 'thi', 'ban', 'bat', 'biet', 'bui',
                  'can', 'cay', 'chi', 'cho', 'con', 'dau', 'dep', 'dia', 'do',
                  'doc', 'don', 'duoc', 'em', 'gan', 'gio', 'hat', 'hieu', 'hoa',
                  'hoi', 'ket', 'khu', 'loai', 'luc', 'luu', 'may', 'mien', 'mo',
                  'mua', 'muon', 'nay', 'ngay', 'nghe', 'ngo', 'nguoi', 'nha',
                  'nhau', 'nhe', 'nhieu', 'nho', 'nom', 'nuoc', 'phai', 'phan',
                  'phong', 'phu', 'qua', 'roi', 'san', 'sau', 'su', 'ta', 'tai',
                  'tam', 'tan', 'tap', 'ten', 'tiet', 'tim', 'tinh', 'toan',
                  'toc', 'toi', 'ton', 'tong', 'trac', 'trang', 'tri', 'troi',
                  'trong', 'trung', 'tuyen', 'ung', 'viec', 'vu', 'xa', 'xay',
                  'xe', 'xin', 'xoa', 'xu', 'xuat', 'xung'}
    words_in_s = set(s.lower().split())
    if words_in_s & viet_words:
        return False
    return True

def translate(en_text, word=''):
    if not en_text or str(en_text).strip() in ('nan', ''):
        return ''
    s = str(en_text).strip()
    if not is_english(s):
        return s  # already Vietnamese
    # Try exact match
    if s in VI_N3:
        return VI_N3[s]
    # Try lowercase
    if s.lower() in VI_N3:
        return VI_N3[s.lower()]
    # Try with semicolon variants (some use ; some use ,)
    alt = s.replace('; ', ', ')
    if alt in VI_N3:
        return VI_N3[alt]
    alt2 = s.replace(', ', '; ')
    if alt2 in VI_N3:
        return VI_N3[alt2]
    # Try first part before semicolon
    first = s.split(';')[0].strip()
    if first in VI_N3:
        return VI_N3[first]
    first2 = s.split(',')[0].strip()
    if first2 in VI_N3:
        return VI_N3[first2]
    return s  # keep original if no translation found

# ── 1. Load all 8 Excel parts ──────────────────────────────────────────────
print("Loading Excel files...")
all_vocab = []
for i in range(1, 9):
    f = INPUT / f"KOERU_JLPT_N3_Vocab_Kanji_OnKun_HanViet_QA_v2_Part0{i}.xlsx"
    df = pd.read_excel(f, sheet_name='Vocab_N3')
    all_vocab.append(df)
combined = pd.concat(all_vocab, ignore_index=True)
print(f"  Total rows: {len(combined)}")

# ── 2. Translate English → Vietnamese ────────────────────────────────────
vi_col = combined['Nghĩa tiếng Việt'].astype(str)
fixed = 0
still_en = 0
for idx, row in combined.iterrows():
    vi = str(row['Nghĩa tiếng Việt'])
    if is_english(vi):
        en_src = str(row['Nghĩa tiếng Anh nguồn']) if pd.notna(row['Nghĩa tiếng Anh nguồn']) else vi
        translated = translate(en_src or vi, str(row['Từ vựng']))
        combined.at[idx, 'Nghĩa tiếng Việt'] = translated
        if not is_english(translated):
            fixed += 1
        else:
            still_en += 1

print(f"  Translated: {fixed}")
print(f"  Still English: {still_en}")

# ── 3. Build word→meaning map from Excel ─────────────────────────────────
word_vi_map = {}
for _, row in combined.iterrows():
    w = str(row['Từ vựng']).strip()
    vi = str(row['Nghĩa tiếng Việt']).strip()
    r = str(row['Cách đọc']).strip()
    if w and vi and not is_english(vi):
        word_vi_map[w] = {'vi': vi, 'reading': r}

print(f"  Excel word→meaning map: {len(word_vi_map)} entries")

# ── 4. Update kanji-data-n3.js words[] from Excel map ─────────────────────
print("\nUpdating kanji-data-n3.js...")
n3_path = JS / 'kanji-data-n3.js'
content = n3_path.read_text(encoding='utf-8')

update_count = 0
for word, data in word_vi_map.items():
    vi = data['vi']
    reading = data['reading']
    # Look for this word in words arrays where meaning might still be English
    # Pattern: {"w":"word","r":"reading","m":"some_english"}
    pattern = r'(\{"w":"' + re.escape(word) + r'","r":"[^"]*","m":")([^"]+)("\})'
    def replacer(m):
        global update_count
        cur_m = m.group(2)
        if is_english(cur_m):
            update_count += 1
            return m.group(1) + vi + m.group(3)
        return m.group(0)
    content = re.sub(pattern, replacer, content)

print(f"  Updated {update_count} word meanings in n3.js")
n3_path.write_text(content, encoding='utf-8')

# Rebuild combined kanji-data.js
parts = []
for lv in ['n5', 'n4', 'n3', 'n2', 'n1']:
    c = (JS / f"kanji-data-{lv}.js").read_text(encoding='utf-8')
    start = c.index('['); end = c.rindex(']') + 1
    inner = c[start+1:end-1].strip().rstrip(',')
    if inner:
        parts.append(inner)
(JS / 'kanji-data.js').write_text(
    'window.KANJI_DATA = [\n' + ',\n'.join(parts) + '\n];\n',
    encoding='utf-8'
)
print("  kanji-data.js rebuilt")

# ── 5. Generate output Excel ───────────────────────────────────────────────
print("\nGenerating output Excel...")

combined['Cần kiểm tra'] = combined['Nghĩa tiếng Việt'].apply(
    lambda x: 'TRUE' if is_english(str(x)) else 'FALSE'
)

wb = Workbook()

# ── Sheet 1: Summary ───────────────────────────────────────────────────────
ws_sum = wb.active
ws_sum.title = 'Summary'
total = len(combined)
translated_count = (combined['Cần kiểm tra'] == 'FALSE').sum()
need_check = (combined['Cần kiểm tra'] == 'TRUE').sum()

header_font = Font(name='Arial', bold=True, size=12)
ws_sum.column_dimensions['A'].width = 30
ws_sum.column_dimensions['B'].width = 20

summary_data = [
    ('File', 'KOERU_JLPT_N3_Updated.xlsx'),
    ('Level', 'N3'),
    ('Tổng từ vựng', total),
    ('Đã dịch sang TV', translated_count),
    ('Cần kiểm tra thêm', need_check),
    ('Tỷ lệ đã dịch', f'{translated_count/total*100:.1f}%'),
    ('Nguồn', '8 file Excel N3 QA v2'),
]
for r, (k, v) in enumerate(summary_data, 1):
    ws_sum[f'A{r}'] = k
    ws_sum[f'B{r}'] = v
    ws_sum[f'A{r}'].font = Font(name='Arial', bold=True)
    ws_sum[f'B{r}'].font = Font(name='Arial')

# ── Sheet 2: Vocab_N3 ──────────────────────────────────────────────────────
ws_v = wb.create_sheet('Vocab_N3')
cols = ['STT', 'JLPT', 'Từ vựng', 'Cách đọc', 'Nghĩa tiếng Việt',
        'Nghĩa tiếng Anh nguồn', 'Cần kiểm tra']
header_fill = PatternFill('solid', start_color='1F5C99')
header_font2 = Font(name='Arial', bold=True, color='FFFFFF', size=10)
thin = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for c, col in enumerate(cols, 1):
    cell = ws_v.cell(1, c, col)
    cell.font = header_font2
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

col_widths = [6, 6, 15, 18, 40, 40, 12]
for c, w in enumerate(col_widths, 1):
    ws_v.column_dimensions[get_column_letter(c)].width = w
ws_v.row_dimensions[1].height = 30
ws_v.freeze_panes = 'A2'

red_fill = PatternFill('solid', start_color='FFE0E0')
yellow_fill = PatternFill('solid', start_color='FFFACD')

for r, row in enumerate(combined.itertuples(index=False), 2):
    need = str(getattr(row, 'Cần kiểm tra', 'FALSE')) == 'TRUE'
    row_fill = red_fill if need else None
    for c, col in enumerate(cols, 1):
        if col == 'STT':
            val = row.STT
        elif col == 'JLPT':
            val = row.JLPT
        elif col == 'Từ vựng':
            val = row._3  # Từ vựng
        elif col == 'Cách đọc':
            val = row._4
        elif col == 'Nghĩa tiếng Việt':
            val = row._5
        elif col == 'Nghĩa tiếng Anh nguồn':
            val = row._6
        else:
            val = row.Cần_kiểm_tra if hasattr(row, 'Cần_kiểm_tra') else getattr(row, 'Cần kiểm tra', '')
        cell = ws_v.cell(r, c, val)
        cell.font = Font(name='Arial', size=10)
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        if row_fill:
            cell.fill = row_fill

print(f"  Vocab sheet: {len(combined)} rows")

wb.save(OUTPUT / 'KOERU_JLPT_N3_Updated.xlsx')
print(f"\n✅ Saved: output/KOERU_JLPT_N3_Updated.xlsx")
print(f"   {translated_count}/{total} translated ({translated_count/total*100:.1f}%)")
print(f"   {need_check} entries still need manual review")
