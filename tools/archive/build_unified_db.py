"""
build_unified_db.py
===================
1. Đọc Excel N5/N4 vocab + kanji
2. Dịch toàn bộ nghĩa tiếng Anh → tiếng Việt
3. Cross-reference với JS database hiện tại
4. Thiết kế lại schema JS (loại bỏ field rác, chuẩn hoá)
5. Xuất: Excel cập nhật + JS database mới
"""
import re, json
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

ROOT   = Path(__file__).parent.parent
JS_DIR = ROOT / "js"
XLS_DIR = ROOT / "input" / "excel"
OUT_DIR = ROOT / "output"
OUT_DIR.mkdir(exist_ok=True)

# ══════════════════════════════════════════════════════════════════════════════
# BẢNG DỊCH TIẾNG ANH → TIẾNG VIỆT (N5 + N4 JLPT vocabulary)
# ══════════════════════════════════════════════════════════════════════════════
VI = {
    # ── Thán từ / Trạng từ / Chức năng ───────────────────────────────────────
    "Ah!, Oh!": "Ồ!, Ôi!",
    "yes": "vâng; có",
    "yes; appears, to be the case": "vâng; đúng vậy",
    "no, not at all": "không; không phải",
    "Hello? (used on the phone)": "A-lô? (trả lời điện thoại)",
    "come now, well": "nào, thôi",
    "please, kindly, by all means": "xin mời; làm ơn",
    "Thank you; somehow; no matter how hard one may try": "cảm ơn; bằng cách nào đó",
    "then, well, so": "thế thì; vậy thì",
    "well, well then": "vậy thì; thôi thì",
    "in that situation, well then...": "trong trường hợp đó, vậy thì...",
    "however; but": "tuy nhiên; nhưng",
    "but, however": "nhưng; tuy nhiên",
    "and; furthermore": "và; hơn nữa",
    "and then, after that": "và rồi; sau đó",
    "very (much), greatly, exceedingly": "rất; vô cùng",
    "always, usually, every time, never (with neg. verb)": "luôn luôn; thường xuyên; mỗi lần",
    "already; again; more": "rồi; thêm nữa; đã",
    "yet, still, besides": "vẫn còn; chưa; ngoài ra",
    "more": "thêm; hơn nữa",
    "immediately, soon": "ngay; ngay lập tức",
    "slowly, at ease": "từ từ; thong thả",
    "frequently, often (much); well, skillfully": "thường; hay; giỏi",
    "straight (ahead), direct": "thẳng; thẳng phía trước",
    "a little, somewhat; just a little, somewhat": "một chút; đôi chút",
    "only ~, just ~, as ~": "chỉ ~; chỉ là ~",
    "et cetera": "vân vân; v.v.",
    "at a time": "mỗi lần; mỗi",
    "past; to exceed, ~ too much": "quá ~; vượt quá",
    "plural suffix": "hậu tố số nhiều",
    "about, toward, approximately (time)": "khoảng; vào khoảng (thời gian)",
    "approximate (quantity)": "khoảng chừng; xấp xỉ (số lượng)",
    "honorable ~ (honorific)": "~ đáng kính; tiền tố kính ngữ",
    "Mr. ~, Ms. ~": "anh/chị/ông/bà ~",
    "feel": "cảm thấy ~",
    # ── Đại từ / Chỉ thị ────────────────────────────────────────────────────
    "you": "bạn; anh; chị",
    "this one": "cái này",
    "that one": "cái đó",
    "that one (over there)": "cái kia",
    "which one": "cái nào",
    "which one, which way": "cái nào; hướng nào",
    "this way (polite)": "phía này (lịch sự)",
    "this person (polite); this way (polite)": "người này (lịch sự); phía này",
    "this person; this direction; this side": "người này; hướng này; phía này",
    "over there": "đằng kia; chỗ kia",
    "this way (polite); over there": "phía này (lịch sự); đằng kia",
    "that place, there; bottom, sole": "nơi đó; ở đó; đế (giày)",
    "over there (polite)": "đằng kia (lịch sự)",
    "there, over there, that place": "ở đó; chỗ đó",
    "here, this place": "ở đây; chỗ này",
    "where, what place": "ở đâu; chỗ nào",
    "which (one) (way); where (polite)": "cái nào; ở đâu (lịch sự)",
    "how, in what way": "như thế nào; bằng cách nào",
    "how, in what way (polite)": "như thế nào (lịch sự)",
    "why, for what reason": "tại sao; vì lý do gì",
    "why (same as どうして)": "tại sao",
    "when": "khi nào",
    "how many, how old": "bao nhiêu; mấy tuổi",
    "how much, how many": "bao nhiêu",
    "what, what kind of": "cái gì; loại nào",
    "such, like this": "như thế này; kiểu này",
    "that, like that": "như vậy; cái đó",
    "this": "cái này",
    "that": "cái đó",
    "that over there; like that, that way; um...": "cái kia; như vậy; ừm...",
    "which": "cái nào",
    "who": "ai",
    # ── Tính từ ──────────────────────────────────────────────────────────────
    "good": "tốt; ngon",
    "bad; unpleasant": "xấu; tệ",
    "terrible (in reference to food), unappetizing, unpleasant (taste)": "dở (đồ ăn); không ngon; khó chịu",
    "blue": "xanh lam; xanh",
    "bright (in reference to personality or weather); cheerful": "sáng; vui vẻ; rạng rỡ",
    "noisy; annoying": "ồn; phiền",
    "boring, dull; insignificant": "nhàm chán; tẻ nhạt; không quan trọng",
    "bustling, busy": "nhộn nhịp; đông đúc",
    "splendid, fine": "tuyệt vời; đẹp đẽ",
    "such, like this (this kind of)": "loại như thế này",
    "there isn't, doesn't have": "không có; không tồn tại",
    "all right; can do": "được rồi; có thể làm",
    # ── Động từ ──────────────────────────────────────────────────────────────
    "to meet, to see": "gặp; gặp gỡ",
    "to open, to become open": "mở; mở ra",
    "to open (v.t.)": "mở (ngoại động từ)",
    "to raise, to lift": "nâng lên; đưa lên",
    "to wash": "rửa; giặt",
    "to play (music), to perform": "chơi (nhạc cụ); biểu diễn",
    "to walk": "đi bộ",
    "to be (a person, animate)": "ở; tồn tại (người, sinh vật)",
    "to say": "nói",
    "to be (inanimate objects)": "có; tồn tại (đồ vật)",
    "to go": "đi",
    "to come": "đến; tới",
    "to return, to go home": "về; trở về",
    "to do": "làm",
    "to do, to try; to wear small items (e.g., necktie, watch, etc.)": "làm; thử; đeo (phụ kiện nhỏ)",
    "to turn on (e.g., a light); to take": "bật (đèn); lấy",
    "to turn off, to switch off": "tắt",
    "to become": "trở thành; trở nên",
    "to be able to (to accomplish)": "có thể; làm được",
    "to eat": "ăn",
    "to drink": "uống",
    "to sleep, to go to bed": "ngủ",
    "to get up; to wake up": "dậy; thức dậy",
    "to put on (items below your waist)": "mặc (quần áo bên dưới)",
    "to wear (on head, body, etc.)": "đội; mặc; đeo",
    "to put on, to wear (items above the waist)": "mặc (áo, quần áo trên)",
    "to wear, to put on (e.g., a hat on the head)": "đội; đeo (mũ)",
    "to put on, to wear (glasses, shoes, etc.)": "đeo (kính); mang (giày)",
    "to buy": "mua",
    "to sell": "bán",
    "to read": "đọc",
    "to write": "viết",
    "to listen, to hear": "nghe",
    "to speak, to talk": "nói; nói chuyện",
    "to understand": "hiểu",
    "to understand, to know": "hiểu; biết",
    "to know": "biết",
    "to think": "nghĩ",
    "to remember, to memorize": "nhớ; ghi nhớ",
    "to forget": "quên",
    "to look, to watch, to see": "nhìn; xem",
    "to show": "chỉ; cho xem",
    "to wait": "đợi; chờ",
    "to stand": "đứng",
    "to sit": "ngồi",
    "to enter": "vào; đi vào",
    "to exit, to leave": "ra; rời đi",
    "to put in; to insert": "bỏ vào; nhét vào",
    "to take out": "lấy ra",
    "to take, to pick up": "lấy; cầm",
    "to give": "cho; tặng",
    "to give (to me)": "cho tôi",
    "to receive": "nhận",
    "to borrow": "mượn; vay",
    "to lend": "cho mượn",
    "to use": "dùng; sử dụng",
    "to make, to create": "làm; tạo ra",
    "to cook": "nấu ăn",
    "to cut": "cắt",
    "to put, to place": "để; đặt",
    "to take a photo": "chụp ảnh",
    "to take a bath": "tắm",
    "to wash (the face, hands)": "rửa (mặt, tay)",
    "to clean (the house)": "dọn dẹp (nhà)",
    "to live, to reside": "sống; cư trú",
    "to work": "làm việc",
    "to study, to learn": "học; nghiên cứu",
    "to teach": "dạy; giảng",
    "to ask": "hỏi",
    "to answer": "trả lời",
    "to explain": "giải thích",
    "to begin, to start": "bắt đầu",
    "to end, to finish": "kết thúc; xong",
    "to continue": "tiếp tục",
    "to stop": "dừng; dừng lại",
    "to run": "chạy",
    "to swim": "bơi",
    "to play (games, sports)": "chơi (trò chơi, thể thao)",
    "to sing": "hát",
    "to dance": "nhảy; khiêu vũ",
    "to draw (a picture)": "vẽ",
    "to write (a letter)": "viết (thư)",
    "to call (by name), to invite": "gọi; mời",
    "to dial/call (e.g., phone); to sit down": "gọi điện (điện thoại); ngồi xuống",
    "to ring (the doorbell)": "bấm chuông",
    "to open (a door, window)": "mở (cửa, cửa sổ)",
    "to close (a door, window)": "đóng (cửa, cửa sổ)",
    "to turn on (a device)": "bật; khởi động",
    "to turn off (a device)": "tắt; tắt máy",
    "to fall (rain, snow)": "rơi; đổ (mưa, tuyết)",
    "to sound, to ring (v.i.)": "kêu; reo (nội động từ)",
    "to cry (of birds, insects)": "kêu (chim, côn trùng)",
    "to travel, to make a trip": "du lịch; đi chơi",
    "to ride (a vehicle)": "đi (xe); lên (xe)",
    "to get off (a vehicle)": "xuống (xe)",
    "to change (train, bus)": "đổi (tàu, xe)",
    "to turn (a direction)": "quẹo; rẽ",
    "to cross (a road, bridge)": "qua (đường, cầu)",
    "to climb": "leo; trèo",
    "to pull": "kéo",
    "to push": "đẩy",
    "to throw": "ném; liệng",
    "to catch": "bắt",
    "to hit": "đánh; gõ",
    "to touch, to feel": "chạm; sờ",
    "to break (something)": "làm vỡ; làm gãy",
    "to break (intransitive)": "vỡ; gãy",
    "to lose (a thing)": "mất (đồ)",
    "to find": "tìm thấy",
    "to look for": "tìm kiếm",
    "to meet, to encounter": "gặp; gặp gỡ",
    "to marry": "kết hôn; lấy chồng/vợ",
    "to be born": "sinh ra; được sinh ra",
    "to die": "chết",
    "to send": "gửi",
    "to receive (receive from a superior)": "nhận (từ người trên)",
    "to take (a photo)": "chụp (ảnh)",
    "it takes (amount of time, money) (v.i.)": "mất (thời gian, tiền)",
    "to step on, to tread on": "đạp; dẫm lên",
    "to steal; to rob": "trộm; cướp",
    "to live (as in 'to be alive')": "sống",
    "to do; to give (to pets, parents, siblings, etc.)": "làm; cho (thú cưng, cha mẹ, anh chị em)",
    # ── Danh từ — Người ─────────────────────────────────────────────────────
    "grandfather, male senior citizen": "ông nội/ngoại; cụ ông",
    "grandmother, female senior-citizen": "bà nội/ngoại; cụ bà",
    "policeman (friendly term)": "chú cảnh sát (thân mật)",
    # ── Danh từ — Đồ vật ─────────────────────────────────────────────────────
    "apartment (abbr.)": "căn hộ; chung cư",
    "cup": "cái cốc; cái ly",
    "bag, basket": "túi; giỏ xách",
    "camera": "máy ảnh",
    "curry (abbr. for curry and rice)": "cà ri (cà ri cơm)",
    "calendar": "lịch; tờ lịch",
    "guitar": "đàn guitar",
    "a class": "lớp học; lớp",
    "gram": "gam",
    "coat; court (e.g., tennis)": "áo khoác; sân (tennis)",
    "coffee": "cà phê",
    "a tumbler; a glass": "ly thủy tinh",
    "elevator": "thang máy",
    "tape": "băng dán; băng",
    "tape recorder": "máy ghi âm băng",
    "table": "bàn (ăn, làm việc)",
    "test": "bài kiểm tra; kỳ thi",
    "(abbr.) department store": "trung tâm thương mại; siêu thị lớn",
    "television, TV": "tivi; truyền hình",
    "door (Western style)": "cánh cửa (kiểu phương Tây)",
    "bathroom; toilet": "nhà vệ sinh; toilet",
    "knife": "con dao",
    "news": "tin tức",
    "tie, necktie": "cà vạt",
    "notebook": "vở; sổ tay",
    "a party": "bữa tiệc; buổi tiệc",
    "bus; bath; bass": "xe buýt; bồn tắm",
    "butter": "bơ",
    "bread": "bánh mì",
    "handkerchief": "khăn tay",
    "film (roll of)": "cuộn phim (chụp ảnh)",
    "swimming pool": "bể bơi; hồ bơi",
    "fork": "nĩa",
    "a page": "trang (sách)",
    "bed": "giường",
    "pet": "thú cưng",
    "pen": "bút",
    "ball-point pen": "bút bi",
    "pocket": "túi (áo, quần)",
    "mailbox; post, position": "hòm thư; bưu điện",
    "button": "cái nút; nút bấm",
    "hotel": "khách sạn",
    "match": "que diêm",
    "meter": "mét",
    "radio": "đài radio",
    "radio cassette player": "máy phát radio cassette",
    "record": "đĩa than; kỷ lục",
    "restaurant": "nhà hàng",
    "shirt (lit: white shirt), business shirt": "áo sơ mi (áo trắng); áo công sở",
    "shirt": "áo sơ mi",
    "shower": "vòi hoa sen; tắm vòi",
    "skirt": "váy",
    "heater (lit: stove)": "lò sưởi",
    "spoon": "cái thìa; cái muỗng",
    "sport(s)": "thể thao",
    "trousers": "quần dài",
    "sweater": "áo len",
    "zero": "số không",
    "taxi": "taxi",
    "tobacco, cigarettes": "thuốc lá",
    "near, close, beside; Japanese traditional buckwheat noodle": "gần; kế bên; mì soba (mì kiều mạch Nhật)",
    "length, height": "chiều dài; chiều cao",
    "home; house; my place": "nhà; nhà tôi",
    "gasoline, petrol": "xăng; nhiên liệu",
    "kilo (kilogram)": "(viết tắt) ki-lô (kilogram)",
    "(abbr.) kilo (kilogram)": "(viết tắt) ki-lô (kilogram)",
    "(abbr.) kilo (kilometer)": "(viết tắt) ki-lô (kilomet)",
    "spring": "mùa xuân",
    "summer": "mùa hè",
    "fall (season)": "mùa thu",
    "winter": "mùa đông",
    "morning": "buổi sáng",
    "blue sky": "bầu trời xanh",
    "soy sauce": "nước tương",
    # ── Danh từ — Thời gian ──────────────────────────────────────────────────
    "year before last": "năm kia",
    # ── Trạng từ ─────────────────────────────────────────────────────────────
    "all, everyone, everybody": "tất cả; mọi người",
    "all right, surely, certainly": "chắc chắn; nhất định",
    "straight, quickly": "thẳng; nhanh chóng",
    "firmly, steady": "vững chắc; kiên định",
    "generally, usually": "thường; nói chung",
    "finally, at last": "cuối cùng; rốt cuộc",
    # ── N4 vocab ─────────────────────────────────────────────────────────────
    "~ district, ~ ward, ~ borough": "quận ~; khu ~",
    "to live (animals), to be alive": "sống (động vật); còn sống",
    "to give (me)": "cho (tôi)",
    "to look after, to take care of": "chăm sóc; trông nom",
    "to move (to a new place)": "chuyển (đến chỗ mới)",
    "to graduate": "tốt nghiệp",
    "to get married": "kết hôn; lấy chồng/vợ",
    "to divorce": "ly hôn",
    "to get divorced": "ly hôn",
    "to be born": "được sinh ra; chào đời",
    "to grow up": "lớn lên",
    "to die": "chết; qua đời",
    "to get sick, to fall ill": "bị ốm; bệnh",
    "to get well, to recover": "hồi phục; khỏi bệnh",
    "to be healed; to be cured": "được chữa khỏi",
    "to rest, to take a day off": "nghỉ; nghỉ ngơi",
    "to be tired": "mệt; mệt mỏi",
    "to exercise, to train": "tập thể dục; luyện tập",
    "to play (musical instrument)": "chơi (nhạc cụ)",
    "to practice": "luyện tập; thực hành",
    "to make a mistake, to be mistaken": "nhầm; mắc lỗi",
    "to scold, to tell off": "mắng; la rầy",
    "to praise": "khen; khen ngợi",
    "to receive (sth from sb)": "nhận được",
    "to send (a letter, package)": "gửi (thư, bưu kiện)",
    "to contact, to get in touch": "liên lạc",
    "to introduce": "giới thiệu",
    "to explain": "giải thích",
    "to report, to tell": "báo cáo; kể lại",
    "to translate": "dịch",
    "to plan": "lập kế hoạch",
    "to decide": "quyết định",
    "to choose": "chọn",
    "to change (plans, etc.)": "thay đổi (kế hoạch...)",
    "to cancel": "huỷ; huỷ bỏ",
    "to reserve, to book": "đặt (chỗ, phòng)",
    "to pay": "trả (tiền)",
    "to earn": "kiếm (tiền)",
    "to spend (money)": "chi tiêu",
    "to save (money)": "tiết kiệm",
    "to lend": "cho mượn",
    "to borrow": "mượn",
    "to repair, to fix": "sửa chữa",
    "to break (a machine, etc.)": "làm hỏng (máy móc)",
    "to use, to utilize": "dùng; sử dụng",
    "to try (doing something)": "thử (làm gì đó)",
    "to seem, to appear": "có vẻ; trông có vẻ",
    "to become, to get (adjective)": "trở thành; trở nên",
    "to increase": "tăng; tăng lên",
    "to decrease": "giảm; giảm xuống",
    "to win": "thắng",
    "to lose (a game)": "thua",
    "to pass (an exam)": "đậu; vượt qua (kỳ thi)",
    "to fail (an exam)": "trượt; rớt (kỳ thi)",
    "to apply for": "nộp đơn; đăng ký",
    "to enter (a school)": "nhập học; vào trường",
    "to leave (a company)": "nghỉ việc; rời công ty",
    "to be promoted": "được thăng chức",
    "to transfer": "chuyển; chuyển công tác",
    "to retire": "nghỉ hưu",
    "to quit, to resign": "nghỉ việc; từ chức",
    "to be absent": "vắng mặt",
    "to be late": "đến trễ; muộn",
    "to hurry": "vội; nhanh lên",
    "to be in time, to make it": "kịp; đến kịp",
    "to go on a trip, to travel": "đi du lịch",
    "to pick up (from a station, etc.)": "đón (ở ga...)",
    "to see off": "tiễn",
    "to pass by, to go past": "đi qua; qua",
    "to arrive": "đến; tới nơi",
    "to depart": "khởi hành; rời",
    "to park (a car)": "đỗ (xe); đậu xe",
    "to get lost": "lạc đường",
    "to help, to assist": "giúp; hỗ trợ",
    "to take care of, to look after": "chăm sóc",
    "to visit (a person)": "thăm; ghé thăm",
    "to pick flowers": "hái hoa",
    "to water (plants)": "tưới (cây)",
    "to grow (plants)": "trồng (cây); trồng",
    "to get, to obtain": "lấy; có được",
    "to drop, to let fall": "đánh rơi; làm rơi",
    "to collect, to gather": "thu thập; gom",
    "to sort, to arrange": "sắp xếp",
    "to pack (baggage)": "đóng gói; xếp hành lý",
    "to clean up, to put away": "dọn dẹp; cất đi",
    "to throw away": "vứt đi; bỏ đi",
    "to recycle": "tái chế",
    "to reserve, to make a reservation": "đặt trước; đặt chỗ",
    "to light (a fire), to turn on": "đốt; thắp; bật",
    "to extinguish, to put out (fire)": "dập tắt; tắt (lửa)",
    "to smoke (cigarettes)": "hút thuốc",
    "to drink alcohol": "uống rượu",
    "to eat too much": "ăn quá nhiều",
    "to skip a meal": "bỏ bữa",
    "to run (a business)": "kinh doanh; điều hành",
    "to sell": "bán",
    "to deliver": "giao hàng",
    "to produce, to manufacture": "sản xuất",
    "to develop": "phát triển",
    "to research, to investigate": "nghiên cứu; điều tra",
    "to present, to show": "trình bày; cho xem",
    "to compare": "so sánh",
    "to check": "kiểm tra",
    "to copy": "sao chép",
    "to print": "in",
    "to save (a file)": "lưu (tệp)",
    "to connect": "kết nối",
    "to count": "đếm",
    "to add, to increase": "thêm; tăng",
    "to subtract, to decrease": "trừ; giảm",
    "to multiply": "nhân",
    "to divide": "chia",
    "to measure": "đo",
    "to weigh": "cân",
    "to warm up": "làm nóng; hâm nóng",
    "to cool down": "làm lạnh; nguội",
    "to light, to illuminate": "chiếu sáng",
    "to close (eyes, mouth)": "nhắm (mắt); ngậm (miệng)",
    "to open (eyes, mouth)": "mở (mắt); há (miệng)",
    "to breathe": "thở",
    "to sneeze": "hắt hơi",
    "to cough": "ho",
    "to hurt, to ache": "đau",
    "to bleed": "chảy máu",
    "to heal, to recover": "lành; hồi phục",
    "to think, to consider": "nghĩ; cân nhắc",
    "to know, to understand": "biết; hiểu",
    "to notice, to realize": "nhận ra; chú ý",
    "to be surprised": "ngạc nhiên",
    "to be interested in": "quan tâm đến; hứng thú",
    "to be worried": "lo lắng; lo ngại",
    "to be glad, to be happy": "vui; hạnh phúc",
    "to be sad": "buồn",
    "to be angry": "tức giận",
    "to be afraid": "sợ hãi",
    "to be bored": "chán",
    "to be relieved": "nhẹ nhõm",
    "to be embarrassed": "ngượng; xấu hổ",
    # ── N4 specific ──────────────────────────────────────────────────────────
    "~ language": "tiếng ~",
    "about (approximately)": "khoảng; vào khoảng",
    "morning (early morning)": "sáng sớm",
    "afternoon; PM": "chiều; buổi chiều",
    "evening": "buổi tối",
    "night": "ban đêm",
    "late at night": "khuya; đêm khuya",
    "noon; midday": "trưa; giữa trưa",
    "midnight": "nửa đêm",
    "day after tomorrow": "ngày kia",
    "day before yesterday": "hôm kia",
    "next week": "tuần sau",
    "last week": "tuần trước",
    "next month": "tháng sau",
    "last month": "tháng trước",
    "next year": "năm sau; năm tới",
    "last year": "năm ngoái",
    "every day": "mỗi ngày; hàng ngày",
    "every week": "mỗi tuần; hàng tuần",
    "every month": "mỗi tháng; hàng tháng",
    "every year": "mỗi năm; hàng năm",
    "recently, these days": "gần đây; dạo này",
    "at that time": "vào lúc đó; khi đó",
    "from now on": "từ nay; từ bây giờ",
    "so far, until now": "cho đến nay; từ trước đến giờ",
    "at first": "lúc đầu; ban đầu",
    "at last, finally": "cuối cùng; rốt cuộc",
    "suddenly": "đột nhiên; bất ngờ",
    "gradually": "dần dần; từ từ",
    "especially": "đặc biệt; nhất là",
    "for example": "ví dụ; chẳng hạn",
    "probably, perhaps": "có lẽ; chắc là",
    "definitely, certainly": "nhất định; chắc chắn",
    "absolutely not, never": "tuyệt đối không; không bao giờ",
    "by all means, please do": "nhất định; xin mời",
    "unfortunately": "tiếc là; đáng tiếc",
    "strangely, oddly": "kỳ lạ thay",
    "actually, in fact": "thực ra; thực tế là",
    "of course, naturally": "tất nhiên; đương nhiên",
    "as expected": "đúng như dự đoán; quả nhiên",
    "on purpose, intentionally": "cố ý; có chủ ý",
    "by accident, coincidentally": "tình cờ; ngẫu nhiên",
    "together, with everyone": "cùng nhau; cùng mọi người",
    "alone, by oneself": "một mình",
    # Thêm từ N4 vocab phổ biến
    "to step on, to tread on": "đạp; dẫm lên",
    "straight, quickly": "thẳng; nhanh chóng",
    "to steal; to rob": "trộm; cướp",
    "generally, usually": "nói chung; thường là",
    "finally, at last": "cuối cùng; rốt cuộc",
    "gasoline, petrol": "xăng",
    "to sound, to ring (v.i.)": "kêu; phát ra tiếng",
    "firmly, steady": "vững chắc; kiên định",
    "to live": "sống",
    "~ area (neighborhood)": "khu vực ~; vùng ~",
    "stairs, staircase": "cầu thang",
    "accident, incident": "tai nạn; sự cố",
    "address": "địa chỉ",
    "age": "tuổi",
    "air": "không khí",
    "air conditioner": "máy lạnh; điều hoà",
    "ambulance": "xe cấp cứu",
    "animal": "động vật",
    "answer": "câu trả lời",
    "area, region": "khu vực; vùng",
    "art, fine art": "nghệ thuật; mỹ thuật",
    "bag (large)": "túi (lớn)",
    "bank": "ngân hàng",
    "bath, bathtub": "bồn tắm",
    "beach": "bãi biển",
    "bicycle": "xe đạp",
    "bill, account": "hoá đơn; tài khoản",
    "blood": "máu",
    "body": "cơ thể",
    "book (counter)": "quyển sách",
    "bookstore": "hiệu sách",
    "boss, superior": "sếp; cấp trên",
    "branch (of a company)": "chi nhánh (công ty)",
    "bridge": "cây cầu",
    "building": "toà nhà",
    "bus stop": "trạm xe buýt",
    "business card": "danh thiếp",
    "button (clothing)": "cái cúc; nút áo",
    "calm, quiet": "bình tĩnh; yên tĩnh",
    "camping": "cắm trại",
    "cap (hat)": "mũ lưỡi trai",
    "card": "thẻ; thiệp",
    "careful, cautious": "cẩn thận; thận trọng",
    "ceiling": "trần nhà",
    "chair": "ghế",
    "change (coins)": "tiền lẻ; tiền thối",
    "channel": "kênh (truyền hình)",
    "character, personality": "tính cách; cá tính",
    "cheap, inexpensive": "rẻ",
    "cherry blossom": "hoa anh đào",
    "chest, breast": "ngực",
    "childhood": "thời thơ ấu",
    "chimney": "ống khói",
    "city hall": "uỷ ban nhân dân thành phố",
    "class, grade": "lớp; học kỳ",
    "cleaning": "dọn dẹp; vệ sinh",
    "clear, fine (weather)": "quang đãng; trời đẹp",
    "clever, smart": "thông minh; lanh lợi",
    "climate, weather": "khí hậu; thời tiết",
    "clothes, clothing": "quần áo",
    "cloud": "mây",
    "cold (illness)": "cảm lạnh; bệnh cảm",
    "college student": "sinh viên đại học",
    "colour": "màu sắc",
    "comfortable, easy": "thoải mái; dễ chịu",
    "company employee": "nhân viên công ty",
    "complain": "phàn nàn",
    "complex, complicated": "phức tạp",
    "concert": "buổi hoà nhạc",
    "condition (health)": "tình trạng sức khoẻ",
    "confirm, check": "xác nhận; kiểm tra",
    "contact": "liên lạc",
    "continue": "tiếp tục",
    "convenient, handy": "tiện lợi",
    "conversation": "cuộc trò chuyện",
    "corner": "góc",
    "country (rural area)": "vùng nông thôn",
    "couple": "cặp đôi",
    "culture": "văn hoá",
    "custom, habit": "thói quen; phong tục",
    "customer, client": "khách hàng",
    "cut, haircut": "cắt; cắt tóc",
    "dangerous": "nguy hiểm",
    "dark": "tối; tối tăm",
    "daughter": "con gái",
    "dear (expensive)": "đắt tiền",
    "delivery": "giao hàng",
    "department (in a company)": "phòng (ban); bộ phận",
    "design": "thiết kế",
    "desk": "bàn làm việc; bàn học",
    "dialogue, conversation": "hội thoại; cuộc đối thoại",
    "diary": "nhật ký",
    "difficult, hard": "khó; khó khăn",
    "direction, way": "hướng; phương hướng",
    "dirty": "bẩn",
    "discount": "giảm giá",
    "disease, illness": "bệnh",
    "dish, cuisine": "món ăn",
    "distance": "khoảng cách",
    "doctor": "bác sĩ",
    "document": "tài liệu; văn bản",
    "dream": "giấc mơ",
    "drink (beverage)": "đồ uống",
    "drive": "lái xe",
    "driving license": "bằng lái xe",
    "easy, simple": "dễ; đơn giản",
    "education": "giáo dục",
    "election": "cuộc bầu cử",
    "electricity": "điện",
    "elementary school": "trường tiểu học",
    "embarrassing, awkward": "ngại; xấu hổ; khó xử",
    "emergency": "tình huống khẩn cấp",
    "energy": "năng lượng",
    "enough, sufficient": "đủ; đủ rồi",
    "entrance, entry": "lối vào; cửa vào",
    "environment": "môi trường",
    "essay": "bài luận",
    "evening party": "buổi tiệc tối",
    "exchange": "trao đổi; đổi",
    "exercise (physical)": "bài tập thể dục",
    "exhibition": "triển lãm",
    "experience": "kinh nghiệm",
    "express train": "tàu tốc hành",
    "eye drops": "thuốc nhỏ mắt",
    "face": "khuôn mặt",
    "factory": "nhà máy; xưởng",
    "famous": "nổi tiếng",
    "fare (transportation)": "giá vé; tiền xe",
    "fault, mistake": "lỗi; sai lầm",
    "favourite": "yêu thích; ưa thích",
    "feature, characteristic": "đặc điểm",
    "fever": "sốt",
    "fire": "lửa; hoả hoạn",
    "first time": "lần đầu tiên",
    "flower viewing": "ngắm hoa",
    "fog": "sương mù",
    "food, diet": "thực phẩm; chế độ ăn",
    "foreigner": "người nước ngoài",
    "forest": "rừng",
    "form, document": "biểu mẫu; tờ đơn",
    "free (no charge)": "miễn phí",
    "free time, leisure": "thời gian rảnh",
    "fresh": "tươi; mới",
    "fruit": "trái cây; hoa quả",
    "full (after eating)": "no",
    "garden": "vườn",
    "gas (fuel)": "khí ga; nhiên liệu",
    "glasses": "kính mắt",
    "gloves": "găng tay",
    "government": "chính phủ",
    "grade, mark": "điểm số",
    "grammar": "ngữ pháp",
    "guest": "khách",
    "guide": "hướng dẫn viên",
    "gym, gymnasium": "phòng tập gym",
    "habit": "thói quen",
    "half": "một nửa",
    "hall, lobby": "sảnh; phòng chờ",
    "health": "sức khoẻ",
    "heart, mind": "trái tim; tâm hồn",
    "heavy": "nặng",
    "help": "sự giúp đỡ",
    "high school": "trường trung học",
    "hill, slope": "đồi; dốc",
    "history": "lịch sử",
    "hobby": "sở thích",
    "hole": "lỗ hổng",
    "homework": "bài tập về nhà",
    "hospital": "bệnh viện",
    "hot spring": "suối nước nóng",
    "house, building": "nhà; toà nhà",
    "hungry": "đói",
    "husband (own)": "chồng (của mình)",
    "idea": "ý tưởng",
    "important": "quan trọng",
    "impossible, no good": "không thể; không được",
    "impression": "ấn tượng",
    "inconvenient": "bất tiện",
    "information": "thông tin",
    "instructions, manual": "hướng dẫn sử dụng",
    "interesting, funny": "thú vị; buồn cười",
    "international": "quốc tế",
    "invitation": "lời mời",
    "island": "hòn đảo",
    "job, work": "công việc",
    "joke": "trò đùa; chuyện hài",
    "journey, trip": "chuyến đi; hành trình",
    "juice": "nước ép; nước trái cây",
    "junior (at school or work)": "đàn em; cấp dưới",
    "key": "chìa khoá",
    "kind (personality)": "tốt bụng; thân thiện",
    "kitchen": "nhà bếp",
    "lake": "hồ",
    "language": "ngôn ngữ; tiếng",
    "large, big": "to; lớn",
    "laundry": "giặt giũ; đồ giặt",
    "leaf": "chiếc lá",
    "lecture": "bài giảng",
    "left (direction)": "bên trái",
    "letter (mail)": "lá thư",
    "library": "thư viện",
    "light (not heavy)": "nhẹ",
    "light (brightness)": "ánh sáng",
    "line (queue)": "hàng; dãy",
    "list": "danh sách",
    "living room": "phòng khách",
    "loud, noisy": "ồn ào",
    "luck": "may mắn",
    "luggage, baggage": "hành lý",
    "magazine": "tạp chí",
    "map": "bản đồ",
    "market": "chợ",
    "marriage": "hôn nhân",
    "meaning": "ý nghĩa",
    "medicine": "thuốc",
    "meeting": "cuộc họp",
    "memory, recollection": "ký ức; kỷ niệm",
    "message": "tin nhắn",
    "mistake, error": "lỗi; sai sót",
    "model": "mô hình; người mẫu",
    "money": "tiền",
    "month": "tháng",
    "mountain": "núi",
    "museum": "bảo tàng",
    "music": "âm nhạc",
    "nationality": "quốc tịch",
    "nature": "thiên nhiên",
    "necessary, needed": "cần thiết",
    "neighborhood": "khu phố; hàng xóm",
    "next": "tiếp theo",
    "nice, pleasant": "dễ chịu; thú vị",
    "night view": "cảnh đêm",
    "noisy": "ồn ào",
    "normal, ordinary": "bình thường",
    "number": "số",
    "nurse": "y tá",
    "object, thing": "đồ vật; thứ",
    "opinion": "ý kiến",
    "order": "đặt hàng; thứ tự",
    "other": "khác; còn lại",
    "outside": "bên ngoài",
    "overtime": "làm thêm giờ",
    "own, oneself": "của mình; bản thân",
    "pain, ache": "cơn đau",
    "painting, drawing": "bức tranh",
    "park (public)": "công viên",
    "parking lot": "bãi đỗ xe",
    "part-time job": "việc làm thêm",
    "passport": "hộ chiếu",
    "patient (hospital)": "bệnh nhân",
    "payment": "thanh toán",
    "people": "người",
    "performance, concert": "buổi biểu diễn",
    "period, era": "thời kỳ; giai đoạn",
    "phone call": "cuộc gọi điện thoại",
    "photo, picture": "ảnh; hình ảnh",
    "platform (train)": "sân ga; cổng",
    "police station": "đồn cảnh sát",
    "politics": "chính trị",
    "pollution": "ô nhiễm",
    "population": "dân số",
    "post office": "bưu điện",
    "power, strength": "sức mạnh; năng lực",
    "present, gift": "quà tặng",
    "price": "giá cả",
    "problem": "vấn đề; bài toán",
    "program, plan": "chương trình; kế hoạch",
    "purpose, goal": "mục đích; mục tiêu",
    "question": "câu hỏi",
    "quiet, calm": "yên tĩnh; bình lặng",
    "rain": "mưa",
    "reason, cause": "lý do; nguyên nhân",
    "receipt": "hoá đơn; biên lai",
    "recipe": "công thức nấu ăn",
    "relationship": "mối quan hệ",
    "repair, fix": "sửa chữa",
    "report": "báo cáo",
    "reservation": "đặt chỗ",
    "review": "ôn tập; đánh giá",
    "reward": "phần thưởng",
    "right (direction)": "bên phải",
    "river": "dòng sông",
    "road, street": "đường phố",
    "room": "căn phòng",
    "rule, regulation": "quy tắc; nội quy",
    "run, operate": "chạy; vận hành",
    "safe, secure": "an toàn",
    "salary": "lương",
    "sale, discount": "khuyến mãi; giảm giá",
    "same": "giống nhau; cùng",
    "scenery, view": "phong cảnh; cảnh quan",
    "schedule": "lịch trình",
    "school": "trường học",
    "science": "khoa học",
    "season": "mùa",
    "seat, place": "chỗ ngồi",
    "senior (at school or work)": "đàn anh; cấp trên",
    "serious, severe": "nghiêm trọng",
    "service": "dịch vụ",
    "shape, form": "hình dạng",
    "shelf": "kệ; giá sách",
    "shopping": "mua sắm",
    "short (not tall)": "thấp; lùn",
    "shy, reserved": "nhút nhát; ngại ngùng",
    "sign, signal": "dấu hiệu; biển báo",
    "skill": "kỹ năng",
    "sky": "bầu trời",
    "sleep": "giấc ngủ",
    "slow": "chậm",
    "small, little": "nhỏ",
    "smell, scent": "mùi; hương",
    "snow": "tuyết",
    "society": "xã hội",
    "sofa": "sofa; ghế sofa",
    "solution": "giải pháp",
    "son": "con trai",
    "song": "bài hát",
    "soup": "súp; canh",
    "south": "phía nam",
    "speed": "tốc độ",
    "stairs": "cầu thang",
    "stamp": "tem thư",
    "station": "ga; nhà ga",
    "strong": "mạnh",
    "student": "học sinh; sinh viên",
    "subject (school)": "môn học",
    "success": "thành công",
    "sugar": "đường",
    "suitable, appropriate": "phù hợp; thích hợp",
    "summer vacation": "kỳ nghỉ hè",
    "supermarket": "siêu thị",
    "swimming": "bơi lội",
    "symptom": "triệu chứng",
    "tall": "cao",
    "temperature": "nhiệt độ",
    "thank you": "cảm ơn",
    "thin": "mỏng; gầy",
    "ticket": "vé",
    "time, hour": "giờ; thời gian",
    "tired, exhausted": "mệt mỏi",
    "today": "hôm nay",
    "tomorrow": "ngày mai",
    "top, upper": "trên; trên cùng",
    "town, city": "thị trấn; thành phố",
    "traffic": "giao thông",
    "train": "tàu hoả; tàu điện",
    "travel, trip": "du lịch; chuyến đi",
    "tree, wood": "cây",
    "trouble, problem": "rắc rối; vấn đề",
    "turn, corner": "chỗ rẽ; góc",
    "umbrella": "ô; dù",
    "uncle": "chú; bác",
    "uncomfortable": "không thoải mái",
    "underground, subway": "tàu điện ngầm",
    "university": "đại học",
    "useful": "hữu ích",
    "vacation, holiday": "kỳ nghỉ; ngày lễ",
    "variety, assortment": "đa dạng; phong phú",
    "vegetables": "rau củ",
    "village": "làng",
    "voice": "giọng nói",
    "volunteer": "tình nguyện viên",
    "walk (take a walk)": "đi dạo",
    "warm": "ấm áp",
    "week": "tuần",
    "weight": "cân nặng",
    "west": "phía tây",
    "window": "cửa sổ",
    "wine": "rượu vang",
    "wish, hope": "ước muốn; hy vọng",
    "wife (own)": "vợ (của mình)",
    "word, vocabulary": "từ; từ vựng",
    "world": "thế giới",
    "worry, anxiety": "lo lắng; băn khoăn",
    "wrong, incorrect": "sai; không đúng",
    "yard, garden": "sân; vườn",
    "year": "năm",
    "young": "trẻ; trẻ trung",
    "zoo": "vườn thú",
}


# ══════════════════════════════════════════════════════════════════════════════
# ĐỌC JS DATABASE → dict word→meaning_vi
# ══════════════════════════════════════════════════════════════════════════════
def load_js_words():
    """Trả về dict: từ → nghĩa tiếng Việt (từ JS database hiện tại)"""
    mapping = {}
    for lv in ['n5', 'n4', 'n3', 'n2', 'n1']:
        txt = (JS_DIR / f"kanji-data-{lv}.js").read_text(encoding='utf-8')
        for block in re.finditer(r'words:\[([^\]]+)\]', txt):
            raw = block.group(1)
            for w in re.finditer(r'"w":"([^"]+)"[^}]*?"m":"([^"]+)"', raw):
                word, meaning = w.groups()
                if word not in mapping:
                    mapping[word] = meaning
    return mapping


def translate(en_text, js_map, word=None):
    """Dịch nghĩa tiếng Anh → tiếng Việt"""
    if not isinstance(en_text, str) or not en_text.strip():
        return ''
    # 1. Nếu từ tồn tại trong JS DB → dùng nghĩa Việt đó
    if word and word in js_map:
        return js_map[word]
    # 2. Tìm trong bảng dịch cứng
    stripped = en_text.strip()
    if stripped in VI:
        return VI[stripped]
    # 3. Tìm kiếm gần đúng (bỏ dấu ngoặc đơn ở cuối)
    base = re.sub(r'\s*\([^)]*\)\s*$', '', stripped)
    if base in VI:
        return VI[base]
    # 4. Giữ nguyên (fallback — sẽ cần review thủ công)
    return en_text


# ══════════════════════════════════════════════════════════════════════════════
# ĐỌC EXCEL VÀ PATCH NGHĨA VIỆT
# ══════════════════════════════════════════════════════════════════════════════
def process_excel(src_path, vocab_sheet, kanji_sheet, level, meaning_col_en):
    """Đọc Excel, thêm cột nghĩa Việt, trả về dataframe đã xử lý"""
    js_map = load_js_words()

    vocab_df  = pd.read_excel(src_path, sheet_name=vocab_sheet)
    kanji_df  = pd.read_excel(src_path, sheet_name=kanji_sheet)
    summary_df = pd.read_excel(src_path, sheet_name='Summary', header=None)

    word_col = 'Từ vựng / Expression' if level == 'N5' else 'Từ vựng'

    # Dịch nghĩa
    vocab_df['Nghĩa VI'] = vocab_df.apply(
        lambda r: translate(r[meaning_col_en], js_map, r[word_col]),
        axis=1
    )

    # Flag còn tiếng Anh
    def still_english(s):
        if not isinstance(s, str): return False
        viet_chars = set('àáảãạăắặằẳẵâầấẩẫậèéẻẽẹêềếểễệìíỉĩịòóỏõọôồốổỗộơờớởỡợùúủũụưừứửữựỳýỷỹỵđ')
        return not any(c in viet_chars for c in s.lower()) and bool(re.search(r'[a-zA-Z]{3,}', s))

    vocab_df['Cần kiểm tra'] = vocab_df['Nghĩa VI'].apply(still_english)

    # Stats
    total   = len(vocab_df)
    vi_done = (~vocab_df['Cần kiểm tra']).sum()
    need_rv = vocab_df['Cần kiểm tra'].sum()
    print(f"\n{level} Vocab: {total} từ | Đã Việt hoá: {vi_done} | Cần review: {need_rv}")

    return vocab_df, kanji_df, summary_df


def write_excel(out_path, level, vocab_df, kanji_df, word_col, reading_col, meaning_col_en):
    """Ghi Excel với format chuẩn"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    total = len(vocab_df)
    vi_done = (~vocab_df['Cần kiểm tra']).sum()

    summary_data = [
        ["KOERU JLPT Database", f"{level} — Vocabulary & Kanji Reference"],
        ["Mô tả", f"Từ vựng + Kanji {level} chuẩn hoá — tiếng Việt, có On/Kun/Hán Việt"],
        ["Chuẩn", "JLPT (tham khảo) + open-anki-jlpt-decks + KANJIDIC"],
        ["Số từ vựng", total],
        ["Đã Việt hoá", int(vi_done)],
        ["Cần review", int(total - vi_done)],
        ["Số kanji", len(kanji_df)],
        ["Cập nhật", "2026"],
    ]
    hdr = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    hdr_fill = PatternFill("solid", start_color="1F3864")
    row_fill = PatternFill("solid", start_color="DCE6F1")

    for i, row in enumerate(summary_data, 1):
        for j, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = Font(name="Arial", size=10, bold=(i == 1))
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if i == 1:
                cell.font = hdr
                cell.fill = hdr_fill
            elif i % 2 == 0:
                cell.fill = row_fill

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 55

    # ── Sheet 2: Vocab ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet(f"Vocab_{level}")

    # Chọn cột xuất
    out_cols = ['STT', word_col, reading_col, 'Nghĩa VI', meaning_col_en,
                'Hán tự trong từ',
                'Hán Việt (theo từng chữ)' if level == 'N4' else 'Hán Việt',
                'Cần kiểm tra', 'Nguồn']
    out_cols = [c for c in out_cols if c in vocab_df.columns]

    # Header style
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    header_fill = PatternFill("solid", start_color="2E4057")
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for j, col in enumerate(out_cols, 1):
        cell = ws2.cell(row=1, column=j, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # Data rows
    for i, (_, row) in enumerate(vocab_df[out_cols].iterrows(), 2):
        needs_review = row.get('Cần kiểm tra', False)
        for j, val in enumerate(row, 1):
            cell = ws2.cell(row=i, column=j, value=val if pd.notna(val) else '')
            cell.font = Font(name="Arial", size=10,
                             color="CC0000" if needs_review and j == 4 else "000000")
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
        if needs_review:
            ws2.cell(row=i, column=4).fill = PatternFill("solid", start_color="FFF2CC")

    # Column widths
    col_widths = {'STT': 6, word_col: 18, reading_col: 16,
                  'Nghĩa VI': 35, meaning_col_en: 35,
                  'Hán tự trong từ': 14,
                  'Hán Việt (theo từng chữ)': 20, 'Hán Việt': 20,
                  'Cần kiểm tra': 14, 'Nguồn': 20}
    for j, col in enumerate(out_cols, 1):
        ws2.column_dimensions[get_column_letter(j)].width = col_widths.get(col, 16)
    ws2.row_dimensions[1].height = 28

    # ── Sheet 3: Kanji ────────────────────────────────────────────────────────
    ws3 = wb.create_sheet(f"Kanji_{level}")
    kanji_out_cols = [c for c in kanji_df.columns if c in kanji_df.columns]

    for j, col in enumerate(kanji_out_cols, 1):
        cell = ws3.cell(row=1, column=j, value=col)
        cell.font = header_font
        cell.fill = PatternFill("solid", start_color="2E4057")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for i, (_, row) in enumerate(kanji_df.iterrows(), 2):
        for j, val in enumerate(row, 1):
            cell = ws3.cell(row=i, column=j, value=val if pd.notna(val) else '')
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="center")
            cell.border = border

    for j, col in enumerate(kanji_out_cols, 1):
        ws3.column_dimensions[get_column_letter(j)].width = 18

    ws3.row_dimensions[1].height = 28

    wb.save(out_path)
    print(f"Đã lưu: {out_path}")


# ══════════════════════════════════════════════════════════════════════════════
# THIẾT KẾ LẠI JS SCHEMA
# ══════════════════════════════════════════════════════════════════════════════
def redesign_js_schema():
    """
    Schema mới:
    - Xoá: meaning_jp, mnemonic, mn_vi (không dùng)
    - Giữ: kanji, hanviet, on, kun, meaning, level, stroke, freq_rank, grade, radical, parts, words
    - Chuẩn hoá words: [{w, r, m}] — m phải là tiếng Việt
    - Sắp xếp field theo thứ tự chuẩn
    """
    FIELD_ORDER = ['kanji', 'hanviet', 'on', 'kun', 'meaning', 'level',
                   'words', 'stroke', 'freq_rank', 'grade', 'radical', 'parts']
    REMOVE_FIELDS = {'meaning_jp', 'meaning_en', 'mnemonic', 'mn_vi'}

    total_removed = 0
    total_entries = 0

    for lv in ['n5', 'n4', 'n3', 'n2', 'n1']:
        path = JS_DIR / f"kanji-data-{lv}.js"
        content = path.read_text(encoding='utf-8')

        # Xoá các field không cần
        for field in REMOVE_FIELDS:
            before = len(content)
            # Match field:"..." or field:["..."] or field:null
            content = re.sub(
                r',?\s*' + field + r'\s*:\s*(?:"[^"]*"|null|\[[^\]]*\])',
                '', content
            )
            removed = before - len(content)
            if removed > 0:
                total_removed += removed
                print(f"  {lv.upper()}: removed field '{field}' ({removed} chars)")

        path.write_text(content, encoding='utf-8')
        count = len(re.findall(r'kanji:"', content))
        total_entries += count

    print(f"\nSchema cleanup: xoá {total_removed} chars | {total_entries} entries còn lại")

    # Rebuild combined
    parts = []
    for lv in ['n5', 'n4', 'n3', 'n2', 'n1']:
        c = (JS_DIR / f"kanji-data-{lv}.js").read_text(encoding='utf-8')
        start = c.index('[')
        end = c.rindex(']') + 1
        inner = c[start+1:end-1].strip().rstrip(',')
        if inner:
            parts.append(inner)

    combined = 'window.KANJI_DATA = [\n' + ',\n'.join(parts) + '\n];\n'
    (JS_DIR / 'kanji-data.js').write_text(combined, encoding='utf-8')
    print("kanji-data.js rebuilt")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == '__main__':
    print("=" * 60)
    print("BƯỚC 1: Xử lý N5 Excel")
    print("=" * 60)
    n5_vocab, n5_kanji, _ = process_excel(
        XLS_DIR / "KOERU_JLPT_N5_Vocab_Kanji_OnKun_HanViet.xlsx",
        vocab_sheet='Vocab_N5',
        kanji_sheet='Kanji_N5',
        level='N5',
        meaning_col_en='Nghĩa nguồn EN'
    )
    write_excel(
        OUT_DIR / "KOERU_JLPT_N5_Updated.xlsx",
        level='N5',
        vocab_df=n5_vocab,
        kanji_df=n5_kanji,
        word_col='Từ vựng / Expression',
        reading_col='Cách đọc / Reading',
        meaning_col_en='Nghĩa nguồn EN'
    )

    print("\n" + "=" * 60)
    print("BƯỚC 2: Xử lý N4 Excel")
    print("=" * 60)
    n4_vocab, n4_kanji, _ = process_excel(
        XLS_DIR / "KOERU_JLPT_N4_Vocab_Kanji_OnKun_HanViet.xlsx",
        vocab_sheet='Vocab_N4',
        kanji_sheet='Kanji_N4',
        level='N4',
        meaning_col_en='Nghĩa EN nguồn'
    )
    write_excel(
        OUT_DIR / "KOERU_JLPT_N4_Updated.xlsx",
        level='N4',
        vocab_df=n4_vocab,
        kanji_df=n4_kanji,
        word_col='Từ vựng',
        reading_col='Cách đọc',
        meaning_col_en='Nghĩa EN nguồn'
    )

    print("\n" + "=" * 60)
    print("BƯỚC 3: Thiết kế lại JS schema")
    print("=" * 60)
    redesign_js_schema()

    print("\n✅ Hoàn tất!")
    print(f"Excel N5: {OUT_DIR}/KOERU_JLPT_N5_Updated.xlsx")
    print(f"Excel N4: {OUT_DIR}/KOERU_JLPT_N4_Updated.xlsx")
