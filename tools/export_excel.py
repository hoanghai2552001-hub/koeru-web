#!/usr/bin/env python3
"""
KOERU — Export to Excel
========================
Dùng: python tools/export_excel.py [output_path]

Tạo file Excel với các sheet:
  • N5 / N4 / N3 / N2 / N1  — kanji info từng level
  • Words                    — từ vựng phẳng (dễ sửa meaning)
  • Radicals                 — bảng bộ thủ

Sau khi sửa trong Excel → chạy: python tools/import_excel.py
"""

import re, json, sys
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment,
                              Border, Side, GradientFill)
from openpyxl.utils import get_column_letter

PROJECT = Path(__file__).parent.parent
JS      = PROJECT / "js"
LEVELS  = ["n5", "n4", "n3", "n2", "n1"]

# ── Màu theo level ────────────────────────────────────────────────────────
LEVEL_COLORS = {
    "n5": ("1a6b3a", "e8f5e9"),   # xanh lá
    "n4": ("1565c0", "e3f2fd"),   # xanh dương
    "n3": ("6a1b9a", "f3e5f5"),   # tím
    "n2": ("e65100", "fff3e0"),   # cam
    "n1": ("b71c1c", "ffebee"),   # đỏ
}
HEADER_FONT_COLOR = "FFFFFF"

# ── Helpers ───────────────────────────────────────────────────────────────

def header_style(ws, row, cols, bg_hex, font_hex="FFFFFF"):
    fill = PatternFill("solid", fgColor=bg_hex)
    font = Font(bold=True, color=font_hex, size=10)
    border = Border(
        bottom=Side(style="medium", color="CCCCCC")
    )
    for col in range(1, cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = fill
        c.font = font
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=False)


def auto_col_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value or "")
                max_len = max(max_len, len(val))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)


def freeze_and_filter(ws, row=1, col=1):
    ws.freeze_panes = ws.cell(row=row + 1, column=col)
    ws.auto_filter.ref = ws.dimensions


def read_level_entries(level: str) -> list[dict]:
    path = JS / f"kanji-data-{level}.js"
    if not path.exists():
        return []
    content = path.read_text(encoding="utf-8")
    entries = []
    for line in content.split("\n"):
        if not line.strip().startswith("{kanji:"):
            continue
        entry = {}
        for field, pat in [
            ("kanji",   r'kanji:"(.)"'),
            ("hanviet", r'hanviet:"(.*?)"'),
            ("on",      r'on:"(.*?)"'),
            ("kun",     r'kun:"(.*?)"'),
            ("meaning", r'meaning:"(.*?)"'),
            ("level",   r'level:"(.*?)"'),
            ("stroke",  r'stroke:(\d+)'),
            ("radical", r'radical:"(.*?)"'),
            ("parts",   r'parts:\[(.*?)\]'),
        ]:
            m = re.search(pat, line)
            entry[field] = m.group(1) if m else ""
        # words[]
        wm = re.search(r'words:\[(.+?)\]', line, re.DOTALL)
        entry["_words_raw"] = wm.group(1) if wm else ""
        entries.append(entry)
    return entries


def parse_words(words_raw: str) -> list[dict]:
    if not words_raw.strip():
        return []
    try:
        return json.loads("[" + words_raw + "]")
    except:
        # Fallback regex
        results = []
        for m in re.finditer(r'\{"w":"(.*?)","r":"(.*?)","m":"(.*?)"\}', words_raw):
            results.append({"w": m.group(1), "r": m.group(2), "m": m.group(3)})
        return results


# ── Sheet builders ────────────────────────────────────────────────────────

def build_level_sheet(wb, level: str, entries: list[dict]):
    color_hdr, color_row = LEVEL_COLORS.get(level, ("333333", "f9f9f9"))
    ws = wb.create_sheet(title=level.upper())

    headers = ["kanji", "hanviet", "on", "kun", "meaning",
               "stroke", "radical", "parts", "level"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    header_style(ws, 1, len(headers), color_hdr)

    row_fill_alt = PatternFill("solid", fgColor=color_row)
    row_fill_def = PatternFill("solid", fgColor="FFFFFF")

    for i, entry in enumerate(entries, 2):
        fill = row_fill_alt if i % 2 == 0 else row_fill_def
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=i, column=col, value=entry.get(h, ""))
            c.fill = fill
            c.alignment = Alignment(vertical="center")
            if h == "kanji":
                c.font = Font(size=14, bold=True)
            elif h == "meaning":
                c.alignment = Alignment(vertical="center", wrap_text=True)

    # Column widths
    widths = {"kanji": 7, "hanviet": 10, "on": 14, "kun": 16,
              "meaning": 30, "stroke": 7, "radical": 30, "parts": 20, "level": 6}
    for col, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(h, 12)

    ws.row_dimensions[1].height = 20
    freeze_and_filter(ws)
    return ws


def build_words_sheet(wb, all_entries: list[tuple[str, list[dict]]]):
    """Sheet Words — flat table: kanji | level | word | reading | meaning"""
    ws = wb.create_sheet(title="Words")
    headers = ["kanji", "level", "word", "reading", "meaning"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    header_style(ws, 1, len(headers), "37474f")

    row = 2
    level_colors = {
        "N5": "e8f5e9", "N4": "e3f2fd", "N3": "f3e5f5",
        "N2": "fff3e0", "N1": "ffebee",
    }
    for kanji_char, level, words in all_entries:
        fill_color = level_colors.get(level, "FFFFFF")
        fill = PatternFill("solid", fgColor=fill_color)
        for w in words:
            ws.cell(row=row, column=1, value=kanji_char).font = Font(size=13, bold=True)
            ws.cell(row=row, column=2, value=level)
            ws.cell(row=row, column=3, value=w.get("w", ""))
            ws.cell(row=row, column=4, value=w.get("r", ""))
            c = ws.cell(row=row, column=5, value=w.get("m", ""))
            c.alignment = Alignment(wrap_text=True)
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = fill
            row += 1

    widths = {"kanji": 7, "level": 6, "word": 14, "reading": 14, "meaning": 40}
    for col, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(h, 14)

    ws.row_dimensions[1].height = 20
    freeze_and_filter(ws)

    # Note for user
    ws.cell(row=1, column=6, value="⚠ Chỉ sửa cột 'meaning'. Không xóa/thêm dòng.")
    ws.cell(row=1, column=6).font = Font(color="CC0000", italic=True, size=9)

    return ws


def build_radicals_sheet(wb):
    """Sheet Radicals — từ kanji-map-data.js"""
    map_path = PROJECT / "kanji-map-data.js"
    if not map_path.exists():
        return

    content = map_path.read_text(encoding="utf-8")
    rad_start = content.find("radicals:")
    if rad_start < 0:
        return

    rad_section = content[rad_start:rad_start + 80000]
    entries = []
    for line in rad_section.split("\n"):
        m = re.match(r'\s+"(.+?)":\s*\{.*?name_ja:"(.*?)".*?meaning:"(.*?)".*?meaning_en:"(.*?)"', line)
        if m:
            entries.append({
                "key": m.group(1),
                "name_ja": m.group(2),
                "meaning_vi": m.group(3),
                "meaning_en": m.group(4),
            })

    ws = wb.create_sheet(title="Radicals")
    headers = ["key", "name_ja", "meaning_vi", "meaning_en"]
    labels  = ["Ký tự", "Tên JP", "Nghĩa VI", "Nghĩa EN"]
    for col, (h, lbl) in enumerate(zip(headers, labels), 1):
        ws.cell(row=1, column=col, value=lbl)
    header_style(ws, 1, len(headers), "4a148c")

    for i, e in enumerate(entries, 2):
        fill = PatternFill("solid", fgColor="f9f0ff" if i % 2 == 0 else "FFFFFF")
        ws.cell(row=i, column=1, value=e["key"]).font   = Font(size=13)
        ws.cell(row=i, column=2, value=e["name_ja"])
        ws.cell(row=i, column=3, value=e["meaning_vi"])
        ws.cell(row=i, column=4, value=e["meaning_en"])
        for col in range(1, 5):
            ws.cell(row=i, column=col).fill = fill

    widths = [8, 16, 25, 25]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    freeze_and_filter(ws)

    note = ws.cell(row=1, column=5, value="⚠ Chỉ sửa meaning_vi / meaning_en")
    note.font = Font(color="CC0000", italic=True, size=9)
    return ws


# ── Cover sheet ───────────────────────────────────────────────────────────

def build_cover(wb):
    ws = wb.create_sheet(title="README", index=0)
    ws.column_dimensions["A"].width = 60
    info = [
        ("KOERU — Kanji Data Editor", True, 14),
        ("", False, 10),
        (f"Xuất ngày: {datetime.now().strftime('%d/%m/%Y %H:%M')}", False, 10),
        ("", False, 10),
        ("CÁCH SỬ DỤNG", True, 11),
        ("1. Sửa meaning trong sheet Words hoặc N5/N4/N3/N2/N1", False, 10),
        ("2. Sửa nghĩa bộ thủ trong sheet Radicals", False, 10),
        ("3. Lưu file Excel", False, 10),
        ("4. Chạy: python tools/import_excel.py", False, 10),
        ("", False, 10),
        ("KHÔNG nên:", True, 11),
        ("• Xóa hoặc thêm dòng trong sheet Words (sẽ gây lỗi sync)", False, 10),
        ("• Sửa cột kanji/word/reading (chỉ dùng để nhận diện)", False, 10),
        ("• Đổi tên sheet", False, 10),
        ("", False, 10),
        ("Để thêm từ mới: dùng python tools/patch.py", False, 10),
    ]
    for i, (text, bold, size) in enumerate(info, 1):
        c = ws.cell(row=i, column=1, value=text)
        c.font = Font(bold=bold, size=size,
                      color="1a237e" if bold else "333333")
    ws.sheet_view.showGridLines = False


# ── Main ──────────────────────────────────────────────────────────────────

def main():
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    out_path = Path(sys.argv[1]) if len(sys.argv) > 1 else \
               PROJECT / f"tools/kanji_data_{ts}.xlsx"

    print(f"\n📊 KOERU Export Excel")
    print(f"   Output: {out_path}\n")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # xóa sheet mặc định

    build_cover(wb)

    # Level sheets + collect words
    all_words: list[tuple[str, str, list]] = []  # (kanji_char, level, words)
    for level in LEVELS:
        entries = read_level_entries(level)
        if not entries:
            print(f"  ⚠  {level.upper()}: không tìm thấy file, bỏ qua")
            continue
        build_level_sheet(wb, level, entries)
        print(f"  ✅  Sheet {level.upper()}: {len(entries)} kanji")

        for entry in entries:
            words = parse_words(entry["_words_raw"])
            if words:
                all_words.append((entry["kanji"], entry["level"], words))

    # Words sheet
    build_words_sheet(wb, all_words)
    total_words = sum(len(w) for _, _, w in all_words)
    print(f"  ✅  Sheet Words: {total_words} từ vựng từ {len(all_words)} kanji")

    # Radicals sheet
    build_radicals_sheet(wb)
    print(f"  ✅  Sheet Radicals: bộ thủ từ kanji-map-data.js")

    wb.save(out_path)
    print(f"\n✅ Đã xuất: {out_path}")
    print(f"   → Sửa xong chạy: python tools/import_excel.py {out_path.name}")


if __name__ == "__main__":
    main()
